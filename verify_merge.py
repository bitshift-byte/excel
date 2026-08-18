#!/usr/bin/env python3
"""校验脚本：逐行逐列对比生成的还原表与真实合并表"""
import openpyxl

GEN = "/Users/machaojin/Downloads/luo-excel-merger/output/还原_06o_0817.xlsx"
REAL = "/Users/machaojin/Downloads/luo-excel-merger/06o 0817(1).xlsx"

def norm(v):
    if v is None or v == "": return ""
    # 数字和文本数字视为等价（真实表内部类型不统一）
    if isinstance(v, (int, float)):
        if isinstance(v, float):
            return str(round(v, 6))  # 四舍五入消除浮点精度误差
        return str(v)
    s = str(v).strip()
    # 处理 #N/A 等错误值
    if s in ("#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!"):
        return ""
    return s

wb_g = openpyxl.load_workbook(GEN, read_only=True, data_only=True)
wb_r = openpyxl.load_workbook(REAL, read_only=True, data_only=True)

total_diffs = 0
total_cells = 0

for sn in ["Sheet4", "Sheet2", "Sheet5", "Sheet3"]:
    ws_g = wb_g[sn]
    ws_r = wb_r[sn]
    
    # Sheet5 特殊：前2行空，表头在第3行
    if sn == "Sheet5":
        start_row = 4  # 数据从第4行开始
    else:
        start_row = 2
    
    # 对比表头
    g_cols = ws_g.max_column
    r_cols = ws_r.max_column
    g_rows = ws_g.max_row
    r_rows = ws_r.max_row
    
    print(f"\n{'='*50}")
    print(f"{sn}: 生成={g_rows}行x{g_cols}列  真实={r_rows}行x{r_cols}列")
    
    # Sheet4 特殊处理：按交货号对齐后逐行对比（行序可能不同）
    if sn == "Sheet4":
        g_rows_iter = list(ws_g.iter_rows(min_row=start_row, values_only=True))
        r_rows_iter = list(ws_r.iter_rows(min_row=start_row, values_only=True))
        max_cols = max(g_cols, r_cols)
        cell_count = 0
        diff_count = 0
        first_diffs = []
        # 交货在第6列（index 5）
        g_indexed = {}
        r_indexed = {}
        for row in g_rows_iter:
            jhd = row[5] if len(row) > 5 else None
            if jhd is not None:
                g_indexed[str(jhd).strip()] = row
        for row in r_rows_iter:
            jhd = row[5] if len(row) > 5 else None
            if jhd is not None:
                r_indexed[str(jhd).strip()] = row
        
        common = set(g_indexed.keys()) & set(r_indexed.keys())
        for jhd in sorted(common):
            g_row = g_indexed[jhd]
            r_row = r_indexed[jhd]
            for ci in range(max_cols):
                g_val = g_row[ci] if ci < len(g_row) else None
                r_val = r_row[ci] if ci < len(r_row) else None
                cell_count += 1
                if norm(g_val) != norm(r_val):
                    diff_count += 1
                    if len(first_diffs) < 5:
                        first_diffs.append(f"  交货{jhd} col{ci+1}: 生成={repr(norm(g_val))[:30]} vs 真实={repr(norm(r_val))[:30]}")
        pct = (1 - diff_count / cell_count) * 100 if cell_count else 100
        print(f"  (按交货号对齐) 总单元格: {cell_count}, 差异: {diff_count}, 一致率: {pct:.2f}%")
        if first_diffs:
            print(f"  前5个差异:")
            for d in first_diffs:
                print(d)
        else:
            print(f"  ✅ 数据值完全一致!")
        total_diffs += diff_count
        total_cells += cell_count
        continue
    
    # 对比数据行
    diff_count = 0
    cell_count = 0
    first_diffs = []
    
    g_rows_iter = list(ws_g.iter_rows(min_row=start_row, values_only=True))
    r_rows_iter = list(ws_r.iter_rows(min_row=start_row, values_only=True))
    
    max_rows = max(len(g_rows_iter), len(r_rows_iter))
    max_cols = max(g_cols, r_cols)
    
    for ri in range(max_rows):
        g_row = g_rows_iter[ri] if ri < len(g_rows_iter) else [None] * max_cols
        r_row = r_rows_iter[ri] if ri < len(r_rows_iter) else [None] * max_cols
        for ci in range(max_cols):
            g_val = g_row[ci] if ci < len(g_row) else None
            r_val = r_row[ci] if ci < len(r_row) else None
            cell_count += 1
            if norm(g_val) != norm(r_val):
                diff_count += 1
                if len(first_diffs) < 5:
                    first_diffs.append(f"  row{ri+start_row} col{ci+1}: 生成={repr(norm(g_val))[:30]} vs 真实={repr(norm(r_val))[:30]}")
    
    total_diffs += diff_count
    total_cells += cell_count
    pct = (1 - diff_count / cell_count) * 100 if cell_count else 100
    print(f"  总单元格: {cell_count}, 差异: {diff_count}, 一致率: {pct:.2f}%")
    if first_diffs:
        print(f"  前5个差异:")
        for d in first_diffs:
            print(d)
    else:
        print(f"  ✅ 完全一致!")

print(f"\n{'='*50}")
print(f"总计: {total_cells} 单元格, {total_diffs} 差异, 一致率: {(1-total_diffs/total_cells)*100:.2f}%")

wb_g.close()
wb_r.close()
