"""
联合利华 Excel 合并筛选系统
- 登录认证（用户名 + 密码，走认证服务）
- 第一步：上传文件，分析所有 Sheet 表头 + 前10行数据
- 第二步：用户纠正表头列名 + 选择参与合并的 Sheet + 选择筛选省份
- 第三步：按列名对齐合并，筛选选中省份的数据，输出 Excel + 预览
"""

import os
import sys
import json
import uuid
import datetime
import secrets
import asyncio
from typing import List, Dict, Optional
from contextlib import asynccontextmanager

from fastapi import FastAPI, UploadFile, File, HTTPException, Form, Request
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from starlette.middleware.base import BaseHTTPMiddleware

from merger import (
    BUILTIN_RULE_ID,
    BUILTIN_RULE,
    get_province_list,
    serialize_cell,
    read_all_sheets,
    SAMPLE_ROWS,
    match_columns_to_rule,
    apply_value_mappings,
    build_pivot_by_delivery,
    build_pivot_by_factory_delivery,
    _format_date_text,
    match_row_province,
    PREVIEW_MAX_ROWS,
    merge_files,
    _base_dir,
    _resource_path,
)

import mail_reader

DATA_DIR = os.path.join(_base_dir(), "data")
os.makedirs(DATA_DIR, exist_ok=True)
MAIL_CONFIG_FILE = os.path.join(DATA_DIR, "mail_config.json")  # 保留用于向后兼容


def get_all_rules() -> list:
    """获取所有规则：内置规则 + 远程规则。
    如果远程规则中已包含内置规则（由认证服务注入），则不再重复添加。"""
    remote = get_remote_rules()
    # 检查远程规则是否已包含内置规则
    has_builtin = any(r.get("id") == "_builtin_default" or r.get("builtin") for r in remote)
    if has_builtin:
        return remote
    return [BUILTIN_RULE] + remote


@asynccontextmanager
async def lifespan(app):
    # 从认证服务获取邮件配置
    app_cfg = fetch_app_config_from_auth_service()
    global APP_CONFIG_CACHE, APP_CONFIG_CACHE_TIME
    APP_CONFIG_CACHE = app_cfg
    import time as _time
    APP_CONFIG_CACHE_TIME = _time.time()
    mail_cfg = app_cfg.get("mail_config", {})
    if mail_cfg.get("enabled"):
        mail_cfg["output_dir"] = OUTPUT_DIR
        mail_cfg["processed_uids_file"] = os.path.join(DATA_DIR, "processed_uids.json")
        mail_reader.start_background(mail_cfg)
    yield
    mail_reader.stop_background()


app = FastAPI(title="Excel 合并筛选系统", lifespan=lifespan)
app.mount("/static", StaticFiles(directory=_resource_path("templates/static")), name="static")

UPLOAD_DIR = os.path.join(_base_dir(), "uploads")
OUTPUT_DIR = os.path.join(_base_dir(), "output")
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)




# ===================== 认证（用户名 + 密码，通过远程认证服务） =====================

def _load_auth_service_url() -> str:
    """
    读取认证服务地址。
    优先级：环境变量 > data/auth_service_url.txt > 默认值
    桌面应用通过编辑 data/auth_service_url.txt 配置服务器地址。
    """
    # 1. 环境变量
    url = os.environ.get("AUTH_SERVICE_URL")
    if url:
        return url.rstrip("/")
    # 2. 配置文件
    url_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "auth_service_url.txt")
    if getattr(sys, "frozen", False):
        if os.name == "nt":
            appdata = os.environ.get("APPDATA")
            base = os.path.join(appdata, "ExcelMerger") if appdata else os.path.dirname(sys.executable)
        else:
            base = os.path.dirname(sys.executable)
        url_file = os.path.join(base, "data", "auth_service_url.txt")
    try:
        if os.path.exists(url_file):
            with open(url_file, "r", encoding="utf-8") as f:
                url = f.read().strip()
                if url:
                    return url.rstrip("/")
    except Exception:
        pass
    # 3. 默认值
    return "http://18.177.82.156:8001"


AUTH_SERVICE_URL = _load_auth_service_url()


def _get_service_token() -> str:
    """获取服务间通信密钥"""
    return os.environ.get("SERVICE_TOKEN", "lx-internal-service-token")


def _get_device_id() -> str:
    """获取本机唯一设备标识（持久化存储在 data 目录）。
    首次调用时生成 UUID 并写入文件，后续读取同一值。"""
    device_file = os.path.join(DATA_DIR, "device_id.txt")
    try:
        if os.path.exists(device_file):
            with open(device_file, "r", encoding="utf-8") as f:
                did = f.read().strip()
                if did:
                    return did
    except Exception:
        pass
    # 生成新的设备 ID
    import uuid as _uuid
    did = _uuid.uuid4().hex
    try:
        with open(device_file, "w", encoding="utf-8") as f:
            f.write(did)
    except Exception:
        pass
    return did


def load_users_from_auth_service() -> dict:
    """从远程认证服务加载用户列表（供 get_current_user 查询）"""
    import httpx
    try:
        resp = httpx.get(
            f"{AUTH_SERVICE_URL}/users",
            headers={"X-Service-Token": _get_service_token()},
            timeout=10,
        )
        if resp.status_code == 200:
            data = resp.json()
            return {u["username"]: u for u in data.get("users", [])}
    except Exception as e:
        print(f"[auth] 无法连接认证服务 ({AUTH_SERVICE_URL}): {e}")
    # 认证服务不可用时返回空字典
    return {}


USERS = load_users_from_auth_service()

# session token → {username} 映射（内存存储，重启失效）
SESSIONS: Dict[str, dict] = {}
SESSION_COOKIE = "nebula_session"
SESSION_MAX_AGE = 86400  # 24h

# session → 上次校验用户状态的时间戳（每 5 分钟向认证服务重新校验一次用户是否仍启用）
SESSION_STATUS_CHECK_INTERVAL = 300  # 5 分钟
SESSION_LAST_CHECK: Dict[str, float] = {}

import time as _time


class AuthMiddleware(BaseHTTPMiddleware):
    """拦截需要认证的路由，未登录重定向到 /login"""

    PUBLIC_PATHS = {"/login", "/api/login", "/favicon.ico"}

    async def dispatch(self, request: Request, call_next):
        path = request.url.path
        # 公开路由直接放行
        if path in self.PUBLIC_PATHS or path.startswith("/static"):
            return await call_next(request)

        # 检查 session
        token = request.cookies.get(SESSION_COOKIE)
        if token and token in SESSIONS:
            username = SESSIONS[token]["username"]
            request.state.username = username

            # 定期向认证服务校验用户是否仍启用（每 5 分钟一次）
            now = _time.time()
            last_check = SESSION_LAST_CHECK.get(token, 0)
            if now - last_check > SESSION_STATUS_CHECK_INTERVAL:
                is_valid, reason = verify_user_status_with_auth_service(username)
                if not is_valid:
                    # 用户已被禁用，清除 session
                    del SESSIONS[token]
                    SESSION_LAST_CHECK.pop(token, None)
                    if path.startswith("/api/"):
                        return JSONResponse({"status": "error", "detail": reason or "账号已被禁用"}, status_code=403)
                    return RedirectResponse("/login", status_code=302)
                # 发送心跳到认证服务，刷新设备绑定活跃时间
                try:
                    import httpx
                    httpx.post(
                        f"{AUTH_SERVICE_URL}/heartbeat",
                        json={"username": username, "device_id": _get_device_id()},
                        headers={"X-Service-Token": _get_service_token()},
                        timeout=5,
                    )
                except Exception:
                    pass
                SESSION_LAST_CHECK[token] = now

            return await call_next(request)

        # API 路由返回 401
        if path.startswith("/api/"):
            return JSONResponse({"status": "error", "detail": "未登录或会话已过期"}, status_code=401)

        # 页面重定向到登录
        return RedirectResponse("/login", status_code=302)


app.add_middleware(AuthMiddleware)


def get_current_user(request: Request) -> Optional[dict]:
    token = request.cookies.get(SESSION_COOKIE)
    if token and token in SESSIONS:
        session = SESSIONS[token]
        username = session["username"]
        user_info = USERS.get(username)
        if user_info:
            return {
                "username": username,
                "name": user_info.get("name", username),
                "role": user_info.get("role", "user"),
                "features": dict(user_info.get("features", {})),
            }
    return None


def require_admin_user(request: Request) -> dict:
    """要求当前用户是管理员，否则抛出 403"""
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="未登录")
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="仅管理员可执行此操作")
    return user


def verify_user_status_with_auth_service(username: str) -> tuple:
    """向认证服务实时校验用户当前是否仍启用。
    返回 (is_valid: bool, reason: str)"""
    import httpx
    try:
        resp = httpx.post(
            f"{AUTH_SERVICE_URL}/verify-user",
            json={"username": username},
            headers={"X-Service-Token": _get_service_token()},
            timeout=10,
        )
        if resp.status_code == 200:
            data = resp.json()
            return (data.get("user", {}).get("enabled", False), "")
        elif resp.status_code == 404:
            return (False, "用户不存在")
    except Exception:
        pass
    # 认证服务不可用时，保守地认为用户仍有效（避免服务故障导致全部登出）
    return (True, "")


def fetch_app_config_from_auth_service() -> dict:
    """从认证服务获取应用配置（邮件配置、功能开关、规则）"""
    import httpx
    try:
        resp = httpx.get(
            f"{AUTH_SERVICE_URL}/app-config",
            headers={"X-Service-Token": _get_service_token()},
            timeout=15,
        )
        if resp.status_code == 200:
            data = resp.json()
            return data.get("config", {})
    except Exception as e:
        print(f"[config] 无法从认证服务获取配置: {e}")
    return {}


# 缓存的应用配置
APP_CONFIG_CACHE: dict = {}
APP_CONFIG_CACHE_TIME: float = 0
APP_CONFIG_CACHE_TTL = 300  # 5 分钟缓存


def get_app_config() -> dict:
    """获取应用配置（带缓存）"""
    global APP_CONFIG_CACHE, APP_CONFIG_CACHE_TIME
    import time as _time
    now = _time.time()
    if not APP_CONFIG_CACHE or (now - APP_CONFIG_CACHE_TIME) > APP_CONFIG_CACHE_TTL:
        APP_CONFIG_CACHE = fetch_app_config_from_auth_service()
        APP_CONFIG_CACHE_TIME = now
    return APP_CONFIG_CACHE


def _get_mail_config() -> dict:
    """获取邮件配置（内部函数）"""
    return get_app_config().get("mail_config", {})


def get_features() -> dict:
    """获取功能开关"""
    return get_app_config().get("features", {})


def get_remote_rules() -> list:
    """获取远程规则列表"""
    return get_app_config().get("rules", [])


# ===================== 路由 =====================

@app.get("/login", response_class=HTMLResponse)
async def login_page():
    with open(_resource_path("templates/login.html"), "r", encoding="utf-8") as f:
        return f.read()


@app.post("/api/login")
async def login_api(request: Request):
    """用户名 + 密码登录：调用认证服务校验"""
    import httpx
    body = await request.json()
    username = body.get("username", "").strip()
    password = body.get("password", "").strip()
    if not username or not password:
        return JSONResponse({"status": "error", "detail": "用户名和密码不能为空"}, status_code=400)

    # 登录前先清除该用户的旧设备绑定（解决同一台电脑重复登录被拒的问题）
    # 这使得"最后登录者获胜"：新登录会替换旧会话，同时保持单设备在线
    try:
        async with httpx.AsyncClient() as client:
            await client.post(
                f"{AUTH_SERVICE_URL}/logout",
                json={"username": username},
                headers={"X-Service-Token": _get_service_token()},
                timeout=5,
            )
    except Exception:
        pass  # 如果清除失败，继续尝试登录

    try:
        async with httpx.AsyncClient() as client:
            resp = await client.post(
                f"{AUTH_SERVICE_URL}/login",
                json={"username": username, "password": password, "device_id": _get_device_id()},
                timeout=15,
            )
    except httpx.ConnectError:
        return JSONResponse({"status": "error", "detail": "认证服务不可用"}, status_code=503)

    data = resp.json()
    if resp.status_code != 200 or data.get("status") != "success":
        return JSONResponse(content=data, status_code=resp.status_code)

    user_info = data["user"]

    # 刷新用户表
    global USERS
    USERS = load_users_from_auth_service()

    # 将用户信息存入 session
    token = secrets.token_hex(16)
    SESSIONS[token] = {"username": user_info["username"]}

    resp = JSONResponse({"status": "success", "user": {
        "username": user_info["username"],
        "name": user_info["name"],
        "role": user_info["role"],
        "features": user_info.get("features", {}),
    }})
    resp.set_cookie(
        SESSION_COOKIE, token,
        max_age=SESSION_MAX_AGE,
        httponly=True,
        samesite="lax",
    )
    return resp


@app.post("/api/logout")
async def logout_api(request: Request):
    token = request.cookies.get(SESSION_COOKIE)
    if token and token in SESSIONS:
        username = SESSIONS[token]["username"]["username"]
        # 通知认证服务清除设备绑定
        import httpx
        try:
            httpx.post(
                f"{AUTH_SERVICE_URL}/logout",
                json={"username": username},
                headers={"X-Service-Token": _get_service_token()},
                timeout=5,
            )
        except Exception:
            pass
        del SESSIONS[token]
    SESSION_LAST_CHECK.pop(token, None)
    resp = JSONResponse({"status": "success"})
    resp.delete_cookie(SESSION_COOKIE)
    return resp


@app.get("/api/me")
async def get_me(request: Request):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "detail": "未登录"}, status_code=401)
    return JSONResponse({"status": "success", "user": {
        "username": user["username"],
        "name": user["name"],
        "role": user["role"],
        "features": user.get("features", {}),
    }})


@app.get("/api/users")
async def list_users(request: Request):
    user = get_current_user(request)
    if not user or user["role"] != "admin":
        raise HTTPException(status_code=403, detail="仅管理员可查看")
    users = [
        {"username": u.get("username", ""), "name": u.get("name", ""), "role": u.get("role", "user")}
        for u in USERS.values()
    ]
    return JSONResponse(content={"status": "success", "users": users})


# ===================== 规则 CRUD =====================

@app.get("/api/rules")
async def list_rules():
    return JSONResponse(content={"status": "success", "rules": get_all_rules()})


# 规则管理已移至认证服务后台管理，桌面应用仅读取使用


@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    with open(_resource_path("templates/index.html"), "r", encoding="utf-8") as f:
        return f.read()


@app.get("/mail", response_class=HTMLResponse)
async def mail_page(request: Request):
    # 邮件捞取已整合为主页 SPA 面板，重定向到首页
    return RedirectResponse("/#mail", status_code=302)


@app.get("/mail/results", response_class=HTMLResponse)
async def mail_results_page(request: Request):
    # 处理结果已整合为主页 SPA 面板，重定向到首页
    return RedirectResponse("/#results", status_code=302)


@app.get("/api/features")
async def get_features_api(request: Request):
    """获取当前用户的功能权限（供前端控制 UI 显示）"""
    user = get_current_user(request)
    if not user:
        return JSONResponse(content={"status": "error", "detail": "未登录"}, status_code=401)
    return JSONResponse(content={"status": "success", "features": user.get("features", {})})


@app.get("/api/regions")
async def get_regions():
    return JSONResponse(content={
        "status": "success",
        "regions": get_province_list(),
    })


@app.post("/api/analyze")
async def analyze_files(files: List[UploadFile] = File(...)):
    if not files:
        raise HTTPException(status_code=400, detail="请上传至少一个文件")

    session_id = str(uuid.uuid4())[:8]
    session_dir = os.path.join(UPLOAD_DIR, session_id)
    os.makedirs(session_dir, exist_ok=True)

    saved_files = []
    for f in files:
        # 统一保存为小写后缀，避免大写扩展名导致后续判断失败
        fname = f.filename
        root, ext = os.path.splitext(fname)
        if ext:
            fname = root + ext.lower()
        save_path = os.path.join(session_dir, fname)
        with open(save_path, "wb") as out:
            out.write(await f.read())
        saved_files.append(fname)

    all_sheets = {}
    all_columns_set = []
    all_columns_seen = set()

    for fname in saved_files:
        filepath = os.path.join(session_dir, fname)
        sheets = read_all_sheets(filepath)
        for sname, (headers, data_rows) in sheets.items():
            sample = []
            for r in data_rows[:SAMPLE_ROWS]:
                sample.append([serialize_cell(c) for c in r])
            key = f"{fname}::{sname}"
            all_sheets[key] = {
                "filename": fname,
                "sheet_name": sname,
                "headers": [str(h) if h else "" for h in headers],
                "row_count": len(data_rows),
                "sample_rows": sample,
            }
            for h in headers:
                hs = str(h) if h else ""
                if hs and hs not in all_columns_seen:
                    all_columns_set.append(hs)
                    all_columns_seen.add(hs)

    # 自动分组：表头完全相同的 sheet
    groups = []
    group_map = {}
    gid = 0
    for key, info in all_sheets.items():
        h_tuple = tuple(info["headers"])
        if h_tuple not in group_map:
            gid += 1
            group_map[h_tuple] = gid
            groups.append({"group_id": gid, "headers": list(info["headers"]), "sheets": []})
        groups[group_map[h_tuple] - 1]["sheets"].append(key)

    return JSONResponse(content={
        "status": "success",
        "session_id": session_id,
        "sheets": list(all_sheets.values()),
        "all_columns": all_columns_set,
        "auto_groups": groups,
        "regions": get_province_list(),
        "rules": get_all_rules(),
    })


@app.post("/api/process")
async def process_files(
    request: Request,
    session_id: str = Form(...),
    mappings: str = Form("{}"),
    selected_sheets: str = Form("[]"),
    provinces: str = Form("[]"),
    rule_id: str = Form(""),
):
    session_dir = os.path.join(UPLOAD_DIR, session_id)
    if not os.path.isdir(session_dir):
        raise HTTPException(status_code=400, detail="会话已过期，请重新上传")

    selected = json.loads(selected_sheets)
    prov_list = json.loads(provinces)
    if not selected:
        raise HTTPException(status_code=400, detail="请至少选择一个 Sheet")

    file_paths = [
        os.path.join(session_dir, fname)
        for fname in os.listdir(session_dir)
        if fname.lower().endswith((".xlsx", ".xls", ".csv", ".tsv"))
    ]

    result = merge_files(
        file_paths=file_paths,
        selected_sheets=selected,
        provinces=prov_list,
        rule_id=rule_id or None,
        output_dir=OUTPUT_DIR,
        output_prefix=("筛选结果" if prov_list else "合并结果"),
        manual_mappings=json.loads(mappings) or None,
    )
    stats = result["stats"]
    stats["selected_sheets"] = len(selected)
    stats["sheet_count"] = 5

    return JSONResponse(content={
        "status": "success",
        "stats": stats,
        "previews": result["previews"],
        "download_url": "/api/download",
    })


@app.get("/api/download")
async def download():
    files = [f for f in os.listdir(OUTPUT_DIR) if f.endswith(".xlsx")]
    if not files:
        raise HTTPException(status_code=404, detail="没有可下载的文件，请先处理")
    files.sort(key=lambda f: os.path.getmtime(os.path.join(OUTPUT_DIR, f)), reverse=True)
    latest = files[0]
    output_path = os.path.join(OUTPUT_DIR, latest)
    return FileResponse(
        output_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=latest,
    )


# ===================== 邮件读取器配置 =====================

@app.get("/api/mail/config")
async def get_mail_config():
    """获取邮件配置（从认证服务获取，桌面应用只读）"""
    cfg = _get_mail_config()
    # 不返回 auth_code 给前端
    safe_cfg = {k: v for k, v in cfg.items() if k != "auth_code"} if cfg else {}
    return JSONResponse(content={"status": "success", "config": safe_cfg, "running": mail_reader.is_running()})


# 邮件配置管理已移至认证服务后台管理，桌面应用仅读取使用


@app.post("/api/mail/start")
async def start_mail(request: Request):
    user = require_admin_user(request)
    # 检查用户是否有邮件功能权限
    if not user.get("features", {}).get("mail_reader", True):
        raise HTTPException(status_code=403, detail="您没有邮件读取功能的权限")
    cfg = _get_mail_config()
    if not cfg or not cfg.get("email"):
        raise HTTPException(status_code=400, detail="邮件配置未设置，请在管理后台配置")
    cfg["output_dir"] = OUTPUT_DIR
    cfg["processed_uids_file"] = os.path.join(DATA_DIR, "processed_uids.json")
    mail_reader.start_background(cfg)
    return JSONResponse(content={"status": "success", "running": mail_reader.is_running()})


@app.post("/api/mail/stop")
async def stop_mail(request: Request):
    require_admin_user(request)
    mail_reader.stop_background()
    return JSONResponse(content={"status": "success", "running": mail_reader.is_running()})


@app.get("/api/mail/status")
async def mail_status():
    return JSONResponse(content={"status": "success", "running": mail_reader.is_running()})


@app.get("/api/mail/logs")
async def mail_logs():
    return JSONResponse(content={"status": "success", "logs": mail_reader.get_logs()})


@app.post("/api/mail/run")
async def run_mail_once(request: Request):
    require_admin_user(request)
    cfg = _get_mail_config()
    if not cfg or not cfg.get("email"):
        raise HTTPException(status_code=400, detail="邮件配置未设置，请在管理后台配置")
    cfg["output_dir"] = OUTPUT_DIR
    cfg["processed_uids_file"] = os.path.join(DATA_DIR, "processed_uids.json")
    try:
        body = await request.json()
    except Exception:
        body = {}
    if body and body.get("date"):
        cfg["date"] = body["date"]
    handled = await asyncio.to_thread(mail_reader.process_once, cfg, True)
    return JSONResponse(content={"status": "success", "handled": handled, "logs": mail_reader.get_logs()})


@app.get("/api/mail/results")
async def mail_results():
    files = []
    if os.path.isdir(OUTPUT_DIR):
        for f in os.listdir(OUTPUT_DIR):
            if f.startswith("邮件合并") and f.endswith(".xlsx"):
                path = os.path.join(OUTPUT_DIR, f)
                st = os.stat(path)
                files.append({
                    "filename": f,
                    "mtime": datetime.datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                    "size": st.st_size,
                })
    files.sort(key=lambda x: x["mtime"], reverse=True)
    return JSONResponse(content={"status": "success", "files": files})


@app.get("/api/mail/results/{filename}")
async def download_mail_result(filename: str):
    safe = os.path.basename(filename)
    path = os.path.join(OUTPUT_DIR, safe)
    if not os.path.isfile(path):
        raise HTTPException(status_code=404, detail="文件不存在")
    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=safe,
    )


@app.get("/api/mail/results/{filename}/preview")
async def preview_mail_result(filename: str):
    safe = os.path.basename(filename)
    path = os.path.join(OUTPUT_DIR, safe)
    if not os.path.isfile(path):
        raise HTTPException(status_code=404, detail="文件不存在")
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    sheets = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 21:
                break
            rows.append([serialize_cell(c) for c in row])
        sheets.append({"sheet_name": sname, "rows": rows, "total_rows": ws.max_row})
    wb.close()
    return JSONResponse(content={"status": "success", "filename": safe, "sheets": sheets})


@app.get("/api/mail/tasks")
async def mail_tasks():
    return JSONResponse(content={"status": "success", "tasks": mail_reader.load_tasks()})


if __name__ == "__main__":
    import sys
    import uvicorn
    import threading
    import time

    # ── 无感升级：启动时应用待更新的版本（仅打包后生效）──
    import updater
    updater.apply_pending_update()

    port = 8000

    print(f"[auth] 认证服务地址: {AUTH_SERVICE_URL}")

    # 桌面窗口模式（开发时直接运行、打包后双击运行都走这里）
    import webview

    # ── 无感升级：后台静默检查并下载新版本（3 秒后执行）──
    def _silent_update_check():
        time.sleep(3)
        try:
            updater.check_and_download_update()
        except Exception as e:
            print(f"[updater] 静默检查失败: {e}")

    threading.Thread(target=_silent_update_check, daemon=True).start()

    class Api:
        def _save(self, src, save_filename):
            import os
            import shutil
            if not os.path.isfile(src):
                return None
            dest = webview.windows[0].create_file_dialog(
                webview.FileDialog.SAVE,
                directory=os.path.expanduser("~"),
                save_filename=save_filename,
            )
            if not dest:
                return None
            if isinstance(dest, (tuple, list)):
                dest = dest[0]
            shutil.copy(src, dest)
            return dest

        def download_file(self, filename):
            import os
            safe = os.path.basename(filename)
            return self._save(os.path.join(OUTPUT_DIR, safe), safe)

        def download_latest(self):
            import os
            files = [f for f in os.listdir(OUTPUT_DIR) if f.endswith(".xlsx")]
            if not files:
                return None
            files.sort(key=lambda f: os.path.getmtime(os.path.join(OUTPUT_DIR, f)), reverse=True)
            return self._save(os.path.join(OUTPUT_DIR, files[0]), files[0])

        def check_update(self):
            """检查是否有新版本（不自动升级，仅返回信息）"""
            import updater
            return updater.check_and_update(auto=False)

        def do_update(self):
            """执行自动升级"""
            import updater
            return updater.check_and_update(auto=True)

        def get_version(self):
            """返回当前版本号"""
            import updater
            return updater.get_current_version()

    # 启动主应用服务
    def _run_server():
        uvicorn.run(app, host="127.0.0.1", port=port, log_level="warning")

    threading.Thread(target=_run_server, daemon=True).start()

    window = webview.create_window(
        "LX捞数据",
        f"http://127.0.0.1:{port}",
        width=1280,
        height=820,
        min_size=(900, 600),
        js_api=Api(),
    )

    # 窗口关闭时通知认证服务清除所有设备绑定，避免下次登录被拒
    def _on_closing():
        for token, session in list(SESSIONS.items()):
            username = session["username"]
            try:
                import httpx
                httpx.post(
                    f"{AUTH_SERVICE_URL}/logout",
                    json={"username": username},
                    headers={"X-Service-Token": _get_service_token()},
                    timeout=3,
                )
            except Exception:
                pass

    window.events.closing += _on_closing

    webview.start()
