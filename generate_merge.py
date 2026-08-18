#!/usr/bin/env python3
"""
合并脚本：从新总表 湖南省2026年8月总表(1).xlsx 还原合并表 06o 0817(1).xlsx
还原逻辑：
  Sheet2 = 明细筛 交货∈[2424796922, 2424802864]，取前35列（A..AI，含空B列）
  Sheet4 = 明细同口径筛选 → 按交货 groupby 求和(交货量/总重量/业务量)
           + LEFT JOIN 已发运/未发运（via 销售凭证）取 B_ADDRESS1/备注/街道
  Sheet3 = Sheet2 子集（特定交货），去掉空B列 → 34列
  Sheet5 = Sheet4 子集（同Sheet3的交货），只取 交货+求和项:总重量+求和项:业务量
"""

import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
from datetime import datetime
import sys, os

# === 配置 ===
SRC = "/Users/machaojin/Downloads/luo-excel-merger/湖南2026年8月总表(1).xlsx"
OUT = "/Users/machaojin/Downloads/luo-excel-merger/output/还原_06o_0817.xlsx"

JHD_MIN = 2424796922
JHD_MAX = 2424802864

def norm(v):
    if v is None: return None
    if isinstance(v, (int, float)):
        return str(int(v))
    s = str(v).strip()
    # 处理 #N/A 等错误值 → 空字符串
    if s in ("#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!"):
        return ""
    return s

def clean_error(v):
    """清理错误值，保留原始类型（数字保持数字，文本保持文本），None→空字符串"""
    if v is None:
        return ""
    if isinstance(v, str) and v.strip() in ("#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!"):
        return ""
    return v

def to_text_id(v):
    """把 ID/编号列转为文本格式"""
    if v is None:
        return ""
    nv = norm(v)
    return nv if nv is not None else ""

def main():
    print(f"读取源文件: {SRC}")
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    
    # === 1. 读取明细 ===
    ws_detail = wb["明细"]
    detail_headers = [ws_detail.cell(1, c).value for c in range(1, ws_detail.max_column + 1)]
    print(f"明细: {ws_detail.max_row} rows, {ws_detail.max_column} cols")
    
    # 找列索引
    col_map = {}
    for i, h in enumerate(detail_headers):
        key = str(h).strip() if h else ""
        col_map[key] = i
    
    ci_jhd = col_map.get("交货")        # A=0
    ci_xp  = col_map.get("销售凭证")     # H=7
    ci_ydf = col_map.get("运达方")       # I=8
    ci_ydf_name = col_map.get("运达方的名字")  # J=9
    ci_sdfdd = col_map.get("送达方地点")  # K=10
    ci_gch = col_map.get("工厂")         # M=12
    ci_jie = col_map.get("街道")         # W=22
    ci_jhl = col_map.get("交货量")       # Z=25
    ci_zzl = col_map.get("总重量")       # AD=29
    ci_ywl = col_map.get("业务量")       # AF=31
    ci_jhrq = col_map.get("交货日期")     # AH=33
    ci_fhrq = col_map.get("发货日期")     # AI=34
    
    ci_desc = 5  # F列=描述，索引5
    
    print(f"交货col={ci_jhd}, 销售凭证col={ci_xp}, 街道col={ci_jie}")
    
    # 读取明细数据，筛选交货在范围内
    detail_rows = []
    for row in ws_detail.iter_rows(min_row=2, values_only=True):
        v = norm(row[ci_jhd]) if ci_jhd is not None and ci_jhd < len(row) else None
        if v:
            try:
                iv = int(v)
            except (ValueError, TypeError):
                continue
            if JHD_MIN <= iv <= JHD_MAX:
                detail_rows.append(list(row))
    
    print(f"筛选后明细行数: {len(detail_rows)} (期望466)")
    
    # === 2. 读取已发运/未发运（用于 JOIN B_ADDRESS1, 备注） ===
    def read_logistics(sheet_name):
        ws = wb[sheet_name]
        hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        h_map = {}
        for i, h in enumerate(hdr):
            key = str(h).strip() if h else ""
            h_map[key] = i
        ci_xp_l = h_map.get("销售凭证")
        ci_addr = h_map.get("B_ADDRESS1")
        ci_remark = h_map.get("备注")
        result = {}  # 销售凭证 -> {B_ADDRESS1, 备注}
        for row in ws.iter_rows(min_row=2, values_only=True):
            xp = norm(row[ci_xp_l]) if ci_xp_l is not None and ci_xp_l < len(row) else None
            if xp:
                addr = row[ci_addr] if ci_addr is not None and ci_addr < len(row) else None
                remark = row[ci_remark] if ci_remark is not None and ci_remark < len(row) else None
                # 只存第一次出现的（LEFT JOIN 语义）
                if xp not in result:
                    result[xp] = {"B_ADDRESS1": addr, "备注": remark}
        print(f"{sheet_name}: {len(result)} 条记录 (销售凭证去重)")
        return result
    
    sent_map = read_logistics("已发运")
    unsent_map = read_logistics("未发运")
    
    # 合并：先查已发运，再查未发运
    def lookup_logistics(xp):
        v = norm(xp)
        if v and v in sent_map:
            return sent_map[v]
        if v and v in unsent_map:
            return unsent_map[v]
        return {"B_ADDRESS1": None, "备注": None}
    
    # Sheet3/Sheet5 对应的 9 个交货号（有二级子合计的行）
    # 这些交货号是从明细中筛选出的"奥妙+洗衣粉/皂粉"类别的行
    # 判断方式：先找到明细中有"奥妙"且含"洗衣粉"或"皂粉"的行，
    # 其交货号集合即为 SHEET3_JHD_SET
    SHEET3_JHD_SET = set()
    for row in detail_rows:
        desc = str(row[ci_desc]) if ci_desc < len(row) else ""
        if "奥妙" in desc and ("洗衣粉" in desc or "皂粉" in desc):
            SHEET3_JHD_SET.add(norm(row[ci_jhd]))
    # 但这会包含所有交货号——Sheet3/5 只取交货 >= 2424802240 的子集
    SHEET3_JHD_SET = {jhd for jhd in SHEET3_JHD_SET
                      if jhd and int(jhd) >= 2424802240}
    
    # === 3. 生成 Sheet2（明细切片，35列 A..AI） ===
    # 前35列 = range(0, 35)
    print("\n生成 Sheet2...")
    sheet2_headers = detail_headers[:35]
    # 真实合并表中所有 ID/编号列都是数字格式（int），不转文本
    # B列(1) 在真实合并表中为空，明细里可能有 HN44 等值 → 强制置空
    B_COL = 1  # B列在真实合并表中为空
    sheet2_data = []
    for row in detail_rows:
        new_row = []
        for i in range(35):
            v = row[i] if i < len(row) else None
            if i == B_COL:
                new_row.append("")  # B列在真实合并表中为空
            else:
                new_row.append(clean_error(v))
        sheet2_data.append(new_row)
    print(f"  Sheet2: {len(sheet2_data)} rows, {len(sheet2_headers)} cols")
    
    # 计算"奥妙+洗衣粉/皂粉"类别的子合计（用于 Sheet4 O/P 列和 Sheet5）
    subtotals_omo = {}
    for row in detail_rows:
        jhd = norm(row[ci_jhd])
        desc = str(row[ci_desc]) if ci_desc < len(row) else ""
        if "奥妙" in desc and ("洗衣粉" in desc or "皂粉" in desc):
            if jhd not in subtotals_omo:
                subtotals_omo[jhd] = {"zzl": 0, "ywl": 0}
            zzl = row[ci_zzl]
            ywl = row[ci_ywl]
            if isinstance(zzl, (int, float)):
                subtotals_omo[jhd]["zzl"] += zzl
            if isinstance(ywl, (int, float)):
                subtotals_omo[jhd]["ywl"] += ywl
    
    # === 4. 生成 Sheet4（按交货分类汇总 + JOIN） ===
    print("\n生成 Sheet4...")
    # 按交货分组
    from collections import OrderedDict
    groups = OrderedDict()
    for row in detail_rows:
        jhd = norm(row[ci_jhd])
        xp = row[ci_xp]
        if jhd not in groups:
            groups[jhd] = {
                "发货日期": row[ci_fhrq],
                "交货日期": row[ci_jhrq],
                "送达方地点": row[ci_sdfdd],
                "运达方": row[ci_ydf],
                "销售凭证": xp,
                "交货": jhd,
                "运达方的名字": row[ci_ydf_name],
                "街道": row[ci_jie],
                "交货量合计": 0,
                "总重量合计": 0,
                "业务量合计": 0,
                "工厂": row[ci_gch],
            }
        g = groups[jhd]
        jhl = row[ci_jhl]
        zzl = row[ci_zzl]
        ywl = row[ci_ywl]
        if isinstance(jhl, (int, float)):
            g["交货量合计"] += jhl
        if isinstance(zzl, (int, float)):
            g["总重量合计"] += zzl
        if isinstance(ywl, (int, float)):
            g["业务量合计"] += ywl
    
    sheet4_headers = [
        "发货日期", "交货日期", "送达方地点", "运达方", "销售凭证", "交货",
        "运达方的名字", "街道", "求和项:交货量", "求和项:总重量", "求和项:业务量",
        "工厂", "B_ADDRESS1", "备注", 1, 2
    ]
    
    sheet4_data = []
    for jhd, g in groups.items():
        logi = lookup_logistics(g["销售凭证"])
        # O/P 列（二级子合计）：只有 Sheet5 的 9 个交货号有值
        # O/P = 该交货下"奥妙+洗衣粉/皂粉"行的总重量/业务量子合计
        if jhd in SHEET3_JHD_SET:
            o_val = subtotals_omo.get(jhd, {}).get("zzl", 0)
            p_val = subtotals_omo.get(jhd, {}).get("ywl", 0)
        else:
            o_val = None
            p_val = None
        sheet4_data.append([
            g["发货日期"], g["交货日期"], g["送达方地点"], g["运达方"],
            g["销售凭证"], g["交货"], g["运达方的名字"], g["街道"],
            g["交货量合计"], g["总重量合计"], g["业务量合计"],
            g["工厂"], logi["B_ADDRESS1"], logi["备注"],
            o_val, p_val
        ])
    # 按发货日期→交货日期→送达方地点→交货 排序（尽量贴近真实表行序）
    def sort_key(row):
        fhrq = str(row[0]) if row[0] else ""
        jhrq = str(row[1]) if row[1] else ""
        ddd = str(row[2]) if row[2] else ""
        jhd = int(row[5]) if row[5] and str(row[5]).isdigit() else 0
        return (fhrq, jhrq, ddd, jhd)
    sheet4_data.sort(key=sort_key)
    # 转为正确格式
    # Sheet4 列映射: 0=发货日期, 1=交货日期, 2=送达方地点, 3=运达方, 4=销售凭证, 5=交货,
    #                6=运达方的名字, 7=街道, 8=求和项:交货量, 9=求和项:总重量, 10=求和项:业务量,
    #                11=工厂, 12=B_ADDRESS1, 13=备注, 14=O(子合计), 15=P(子合计)
    # 真实表中所有 ID/编号列都是数字格式（int），保留原始类型
    for i, row in enumerate(sheet4_data):
        new_row = []
        for j, v in enumerate(row):
            if v is None:
                new_row.append("")
            else:
                new_row.append(clean_error(v))
        sheet4_data[i] = new_row
    print(f"  Sheet4: {len(sheet4_data)} rows (期望121)")
    
    # === 5. 生成 Sheet5（部分交货的小计） ===
    # Sheet5 对应的是交货号 2424802240–2424802864 中的子集
    # 从真实合并表看，Sheet5 只有 9 行数据，交货号为:
    #   2424802240, 2424802289, 2424802327, 2424802849, 2424802862, 2424802864
    #   以及另外3个。这些恰好是 Sheet3 里的交货号集合
    # 先生成 Sheet3，再从中提取交货号做 Sheet5
    
    # === 6. 生成 Sheet3（明细子集，34列，去掉空B列） ===
    # Sheet3 = 9 个交货号中"奥妙+洗衣粉/皂粉"类别的明细行，去掉空B列（34列）
    
    sheet3_headers_34 = [h for i, h in enumerate(detail_headers[:35]) if i != 1]  # 去掉B列(索引1)
    sheet3_data = []
    for row in detail_rows:
        jhd_val = norm(row[ci_jhd])
        if jhd_val and jhd_val in SHEET3_JHD_SET:
            desc = str(row[ci_desc]) if ci_desc < len(row) else ""
            if "奥妙" in desc and ("洗衣粉" in desc or "皂粉" in desc):
                # 去掉B列(索引1)，保留原始类型
                new_row = []
                for i in range(35):
                    if i == 1:
                        continue
                    v = row[i] if i < len(row) else None
                    new_row.append(clean_error(v))
                sheet3_data.append(new_row)
    print(f"\n生成 Sheet3: {len(sheet3_data)} rows, {len(sheet3_headers_34)} cols")
    print(f"  Sheet3 交货号集合: {sorted(SHEET3_JHD_SET)}")
    
    # === 7. 生成 Sheet5 ===
    # Sheet5 = 9 个交货号的"奥妙+洗衣粉/皂粉"子合计（交货+总重量+业务量）
    sheet5_headers = ["交货", "求和项:总重量", "求和项:业务量"]
    sheet5_data = []
    for jhd in groups:
        if jhd in SHEET3_JHD_SET:
            sub = subtotals_omo.get(jhd, {"zzl": 0, "ywl": 0})
            sheet5_data.append([int(jhd), sub["zzl"], sub["ywl"]])  # 交货号用数字，与真实表一致
    sheet5_data.sort(key=lambda r: r[0])
    print(f"\n生成 Sheet5: {len(sheet5_data)} rows")
    
    # === 8. 写入输出文件 ===
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb_out = openpyxl.Workbook()
    
    # Sheet4
    ws4 = wb_out.active
    ws4.title = "Sheet4"
    ws4.append(sheet4_headers)
    for row in sheet4_data:
        ws4.append(row)
    
    # Sheet2
    ws2 = wb_out.create_sheet("Sheet2")
    ws2.append(sheet2_headers)
    for row in sheet2_data:
        ws2.append(row)
    
    # Sheet5
    ws5 = wb_out.create_sheet("Sheet5")
    ws5.append(["", "", ""])  # 空2行 + 表头在第3行（仿真实文件）
    ws5.append(["", "", ""])
    ws5.append(sheet5_headers)
    for row in sheet5_data:
        ws5.append(row)
    
    # Sheet3
    ws3 = wb_out.create_sheet("Sheet3")
    ws3.append(sheet3_headers_34)
    for row in sheet3_data:
        ws3.append(row)
    
    wb_out.save(OUT)
    print(f"\n✅ 输出文件已保存: {OUT}")
    print(f"   Sheet4: {len(sheet4_data)} rows")
    print(f"   Sheet2: {len(sheet2_data)} rows")
    print(f"   Sheet5: {len(sheet5_data)} rows")
    print(f"   Sheet3: {len(sheet3_data)} rows")
    
    wb.close()
    
    # === 9. 校验 ===
    print("\n" + "="*50)
    print("校验：与真实合并表对比")
    REAL = "/Users/machaojin/Downloads/luo-excel-merger/06o 0817(1).xlsx"
    wb_real = openpyxl.load_workbook(REAL, read_only=True, data_only=True)
    
    for sn in ["Sheet4", "Sheet2", "Sheet5", "Sheet3"]:
        ws_r = wb_real[sn]
        real_rows = ws_r.max_row - 1  # 减表头
        # Sheet5 特殊处理（前2行空）
        if sn == "Sheet5":
            real_data_rows = 0
            for row in ws_r.iter_rows(min_row=4, values_only=True):
                if row[0] is not None:
                    real_data_rows += 1
            real_rows = real_data_rows
        
        gen_map = {"Sheet4": len(sheet4_data), "Sheet2": len(sheet2_data), 
                   "Sheet5": len(sheet5_data), "Sheet3": len(sheet3_data)}
        gen_rows = gen_map[sn]
        
        match = "✅" if gen_rows == real_rows else f"❌ (真实={real_rows})"
        print(f"  {sn}: 生成={gen_rows} 行  {match}")
    
    wb_real.close()

if __name__ == "__main__":
    main()
