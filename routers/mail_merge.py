"""
邮件合并路由
- 从邮件捞取产物中选一个 + 上传总表 → 将每日数据追加到总表格式 → 输出带日期的总表
"""
import os
import shutil
import uuid
import datetime
from typing import List
from fastapi import APIRouter, Request, UploadFile, File, HTTPException, Form
from fastapi.responses import JSONResponse, FileResponse

import openpyxl

import config
import auth
from merger import (
    fromtimestamp_cn,
    read_all_sheets,
    match_columns_to_rule,
    serialize_cell,
    merge_files,
    merge_mail_into_master,
    load_rules,
    select_source_sheets,
    SAMPLE_ROWS,
)

router = APIRouter(prefix="/api/mail-merge", tags=["mail-merge"])


# ===================== 列出邮件捞取产物 =====================

@router.get("/mail-results")
async def list_mail_results(request: Request):
    """列出 output 目录中邮件捞取的产物文件"""
    if not auth.get_current_user(request):
        raise HTTPException(status_code=401, detail="未登录或登录已过期")

    files = []
    if os.path.isdir(config.OUTPUT_DIR):
        for f in os.listdir(config.OUTPUT_DIR):
            if f.startswith("邮件合并") and not f.startswith("邮件合并结果") and not f.startswith("湖南") and f.endswith(".xlsx"):
                path = os.path.join(config.OUTPUT_DIR, f)
                st = os.stat(path)

                sheet_info = []
                try:
                    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                    for sn in wb.sheetnames:
                        ws = wb[sn]
                        sheet_info.append({
                            "name": sn,
                            "rows": ws.max_row - 1 if ws.max_row > 0 else 0,
                            "cols": ws.max_column,
                        })
                    wb.close()
                except Exception:
                    pass

                files.append({
                    "filename": f,
                    "mtime": fromtimestamp_cn(st.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                    "size": st.st_size,
                    "sheets": sheet_info,
                })
    files.sort(key=lambda x: x["mtime"], reverse=True)
    return JSONResponse(content={"status": "success", "files": files})


# ===================== 选邮件产物 + 上传总表 → 追加到总表 =====================

@router.post("/run")
async def run_merge(
    request: Request,
    files: List[UploadFile] = File(default=[]),
    mail_filename: str = Form(default=""),
    delivery_min: str = Form(default=""),
    delivery_max: str = Form(default=""),
):
    """选邮件产物 + 上传总表 → 将每日数据追加到总表格式

    流程：
    1. mail_filename: 从邮件捞取产物中选的文件（output 目录中）
    2. files: 用户上传的总表文件（湖南2026年8月总表.xlsx）

    将邮件捞取的筛选数据中新增的交货号追加到总表的「明细」sheet，
    保留总表原始格式（已发运/未发运/明细/客户信息/组套/Sheet5）。
    输出文件名带日期。
    """
    if not auth.get_current_user(request):
        raise HTTPException(status_code=401, detail="未登录或登录已过期")

    session_id = str(uuid.uuid4())[:8]
    session_dir = os.path.join(config.UPLOAD_DIR, session_id)
    os.makedirs(session_dir, exist_ok=True)

    # 1. 复制选中的邮件产物
    mail_path = None
    source_desc = []
    if mail_filename:
        mail_src = os.path.join(config.OUTPUT_DIR, mail_filename)
        if not os.path.isfile(mail_src):
            raise HTTPException(status_code=400, detail=f"邮件产物不存在: {mail_filename}")
        mail_path = os.path.join(session_dir, mail_filename)
        shutil.copy2(mail_src, mail_path)
        source_desc.append({"name": mail_filename, "source": "邮件捞取产物"})

    # 2. 保存上传的总表
    master_path = None
    for f in files:
        if not f.filename:
            continue
        fname = f.filename
        root, ext = os.path.splitext(fname)
        if ext:
            fname = root + ext.lower()
        save_path = os.path.join(session_dir, fname)
        with open(save_path, "wb") as out:
            content_bytes = await f.read()
            out.write(content_bytes)
        master_path = save_path
        source_desc.append({"name": fname, "source": "上传总表"})

    if not mail_path:
        raise HTTPException(status_code=400, detail="请选择邮件捞取产物")
    if not master_path:
        raise HTTPException(status_code=400, detail="请上传总表文件")

    # 生成带日期的输出文件名
    today_str = fromtimestamp_cn(datetime.datetime.now().timestamp()).strftime("%Y%m%d")
    output_filename = f"湖南2026年8月总表_{today_str}.xlsx"
    output_path = os.path.join(config.OUTPUT_DIR, output_filename)

    # 执行合并：将邮件数据追加到总表
    result = merge_mail_into_master(master_path, mail_path, output_path)

    stats = {
        "appended_count": result["appended_count"],
        "total_in_detail": result["total_in_detail"],
        "appended_weifayun_count": result.get("appended_weifayun_count", 0),
        "source_files": source_desc,
        "output_filename": output_filename,
    }

    return JSONResponse(content={
        "status": "success",
        "stats": stats,
        "download_url": "/api/mail-merge/download",
    })


# ===================== 下载结果 =====================

@router.get("/download")
async def download(request: Request):
    """下载最近的合并结果（总表格式，带日期）"""
    if not auth.get_current_user(request):
        raise HTTPException(status_code=401, detail="未登录或登录已过期")
    # 优先找总表格式的输出文件
    files = [f for f in os.listdir(config.OUTPUT_DIR)
             if f.startswith("湖南2026年8月总表_") and f.endswith(".xlsx")]
    if not files:
        files = [f for f in os.listdir(config.OUTPUT_DIR)
                 if f.startswith("邮件合并结果") and f.endswith(".xlsx")]
    if not files:
        files = [f for f in os.listdir(config.OUTPUT_DIR) if f.endswith(".xlsx")]
    if not files:
        raise HTTPException(status_code=404, detail="没有可下载的文件")
    files.sort(key=lambda f: os.path.getmtime(os.path.join(config.OUTPUT_DIR, f)), reverse=True)
    latest = files[0]
    return FileResponse(
        os.path.join(config.OUTPUT_DIR, latest),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=latest,
    )
