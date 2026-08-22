"""
邮件读取器路由
- 邮件配置、手动执行、结果列表/下载/预览、任务列表
"""

import os
import asyncio
import datetime
from fastapi import APIRouter, Request, HTTPException
from fastapi.responses import JSONResponse, FileResponse

import config
import auth
import mail_reader
from merger import fromtimestamp_cn

router = APIRouter(prefix="/api/mail", tags=["mail"])


@router.get("/config")
async def get_mail_config(request: Request):
    """获取邮件配置（从认证服务获取，桌面应用只读）。
    provinces 字段按用户隔离：优先使用管理员分配给该用户的省份，
    如果用户没有分配省份，则回退到全局 mail_config 中的省份。"""
    user = auth.get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="未登录或登录已过期")
    cfg = auth.get_mail_config()
    safe_cfg = {k: v for k, v in cfg.items() if k != "auth_code"} if cfg else {}
    # 按用户隔离省份
    user = auth.get_current_user(request)
    username = user.get("username", "") if user else ""
    user_provinces = auth.get_user_provinces(username) if username else []
    if user_provinces:
        safe_cfg["provinces"] = user_provinces
    elif "provinces" not in safe_cfg:
        safe_cfg["provinces"] = []
    return JSONResponse(content={"status": "success", "config": safe_cfg})



@router.post("/run")
async def run_mail_once(request: Request):
    auth.require_admin_user(request)
    cfg = auth.get_mail_config()
    if not cfg or not cfg.get("email"):
        raise HTTPException(status_code=400, detail="邮件配置未设置，请在管理后台配置")
    cfg["output_dir"] = config.OUTPUT_DIR
    cfg["processed_uids_file"] = None  # 使用数据库存储
    try:
        body = await request.json()
    except Exception:
        body = {}
    if body and body.get("date"):
        cfg["date"] = body["date"]
    handled = await asyncio.to_thread(mail_reader.process_once, cfg, True)
    return JSONResponse(content={"status": "success", "handled": handled, "logs": mail_reader.get_logs()})


@router.get("/results")
async def mail_results(request: Request):
    if not auth.get_current_user(request):
        raise HTTPException(status_code=401, detail="未登录或登录已过期")
    files = []
    if os.path.isdir(config.OUTPUT_DIR):
        for f in os.listdir(config.OUTPUT_DIR):
            if f.startswith("邮件合并") and f.endswith(".xlsx"):
                path = os.path.join(config.OUTPUT_DIR, f)
                st = os.stat(path)
                files.append({
                    "filename": f,
                    "mtime": fromtimestamp_cn(st.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                    "size": st.st_size,
                })
    files.sort(key=lambda x: x["mtime"], reverse=True)
    return JSONResponse(content={"status": "success", "files": files})


@router.get("/results/{filename}")
async def download_mail_result(filename: str, request: Request):
    if not auth.get_current_user(request):
        raise HTTPException(status_code=401, detail="未登录或登录已过期")
    safe = os.path.basename(filename)
    path = os.path.join(config.OUTPUT_DIR, safe)
    if not os.path.isfile(path):
        raise HTTPException(status_code=404, detail="文件不存在")
    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=safe,
    )


@router.get("/results/{filename}/preview")
async def preview_mail_result(filename: str, request: Request):
    if not auth.get_current_user(request):
        raise HTTPException(status_code=401, detail="未登录或登录已过期")
    safe = os.path.basename(filename)
    path = os.path.join(config.OUTPUT_DIR, safe)
    if not os.path.isfile(path):
        raise HTTPException(status_code=404, detail="文件不存在")
    import openpyxl
    from merger import serialize_cell
    wb = openpyxl.load_workbook(path, read_only=True)
    sheets = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        all_rows = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append([serialize_cell(c) for c in row])
        total = len(all_rows)
        if total <= 31:
            # 数据量不大，全部返回
            head_rows = all_rows
            tail_rows = []
            tail_start = None
        else:
            # 前 21 行（含表头）+ 后 10 行
            head_rows = all_rows[:21]
            tail_rows = all_rows[-10:]
            tail_start = total - 10
        sheets.append({
            "sheet_name": sname,
            "rows": head_rows,
            "tail_rows": tail_rows,
            "tail_start": tail_start,
            "total_rows": total,
        })
    wb.close()
    return JSONResponse(content={"status": "success", "filename": safe, "sheets": sheets})


@router.get("/tasks")
async def mail_tasks(request: Request):
    if not auth.get_current_user(request):
        raise HTTPException(status_code=401, detail="未登录或登录已过期")
    return JSONResponse(content={"status": "success", "tasks": mail_reader.load_tasks()})

