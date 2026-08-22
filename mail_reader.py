"""邮件读取器：从 126 邮箱读取 Excel 附件，自动合并筛选。"""
import os
import re
import json
import time
import base64
import threading
import imaplib
import email
import datetime
import tempfile
from email.header import decode_header
from typing import List, Set
from collections import deque

from merger import merge_files, _base_dir, now_cn, fromtimestamp_cn

try:
    import database as _db
except ImportError:
    _db = None


# ---------- 工厂标识识别（STDD: 纯逻辑，不依赖外部状态） ----------

# 工厂关键字 → 目标工厂值
# 按邮件中出现的标识（主题/附件名/正文）自动识别实际发货工厂
# 只匹配具体工厂编号，不匹配仓库名称缩写（RDC1/RDC2等）
FACTORY_KEYWORD_MAP = {
    "701": "701",
    "801": "801",
    "901": "901",
    "YG": "YG",
}


def _strip_emails(text: str) -> str:
    """移除文本中的邮箱地址，避免邮箱名中的数字被误匹配为工厂关键字。
    例如 hf901sys@pgl-world.com 中的 901 不应被识别为工厂901。
    """
    return re.sub(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', '', text)


def _factory_keyword_match(keyword: str, text: str) -> bool:
    """检查文本中是否包含工厂关键字（独立词匹配）。
    要求关键字前后不是数字也不是小数点，避免从数字中间截取
    （如 23.901 吨里的 901、293023.90154.452 里的 901 被误匹配）。
    """
    pat = r'(?<![0-9.])' + re.escape(keyword) + r'(?![0-9.])'
    return bool(re.search(pat, text, re.IGNORECASE))


def detect_factory_override(subject: str, attachment_names: list, body_text: str) -> str:
    """从邮件主题、附件名、正文中识别工厂关键字，返回目标工厂值。

    匹配优先级：
    1. 附件名中的工厂关键字（最精确，单附件级别）
    2. 主题中的工厂关键字
    3. 正文中的工厂关键字（最宽松，可能误匹配）

    工厂关键字优先级：701 > 801 > 901 > YG
    如果未识别到，返回空字符串。
    """
    # 层级1：附件名（最精确）
    att_combined = " ".join(attachment_names or []).upper()
    for keyword, target in FACTORY_KEYWORD_MAP.items():
        if _factory_keyword_match(keyword, att_combined):
            return target

    # 层级2：主题
    if subject:
        subj_upper = subject.upper()
        for keyword, target in FACTORY_KEYWORD_MAP.items():
            if _factory_keyword_match(keyword, subj_upper):
                return target

    # 层级3：正文（最宽松，只取前2000字符）
    # 注意：先移除邮箱地址，避免 hf901sys@xxx 中的 901 被误匹配
    if body_text:
        body_clean = _strip_emails(body_text[:2000])
        body_upper = body_clean.upper()
        for keyword, target in FACTORY_KEYWORD_MAP.items():
            if _factory_keyword_match(keyword, body_upper):
                return target

    return ""


def extract_mail_body(msg) -> str:
    """从 email.message.Message 中提取纯文本正文"""
    body = ""
    if msg.is_multipart():
        for part in msg.walk():
            ct = part.get_content_type()
            if ct == "text/plain":
                charset = part.get_content_charset() or "utf-8"
                payload = part.get_payload(decode=True)
                if payload:
                    try:
                        body += payload.decode(charset, errors="replace")
                    except (LookupError, UnicodeDecodeError):
                        body += payload.decode("utf-8", errors="replace")
    else:
        ct = msg.get_content_type()
        if ct == "text/plain":
            charset = msg.get_content_charset() or "utf-8"
            payload = msg.get_payload(decode=True)
            if payload:
                try:
                    body = payload.decode(charset, errors="replace")
                except (LookupError, UnicodeDecodeError):
                    body = payload.decode("utf-8", errors="replace")
    return body


def apply_factory_override(output_path: str, file_factory_map: dict, collected_dir: str) -> int:
    """对合并产物做后处理：将指定附件来源的数据行的「工厂」字段覆盖为目标值。

    file_factory_map: {附件原始文件名: 目标工厂值}
    collected_dir: 临时文件目录（下载的附件在这里）
    返回修改的行数。

    逻辑：
    1. 从临时目录中读取每个附件的交货号集合
    2. 打开输出文件，遍历「全量数据」和「筛选数据」sheet
    3. 根据交货号匹配行，覆盖工厂字段
    """
    if not file_factory_map:
        return 0

    # 步骤1：从临时文件中收集每个附件的交货号 → 目标工厂值映射
    import openpyxl
    jh_to_factory = {}  # 交货号 → 目标工厂值

    for att_name, target_factory in file_factory_map.items():
        # 在 collected_dir 中查找匹配的文件（文件名带序号前缀 N_）
        found = False
        for fname in os.listdir(collected_dir):
            # 去掉序号前缀后比较
            clean_name = re.sub(r"^\d+_", "", fname)
            if clean_name == att_name or fname.endswith("_" + att_name) or att_name in fname:
                fpath = os.path.join(collected_dir, fname)
                try:
                    sheets = read_all_sheets_for_override(fpath)
                    for sname, (headers, data_rows) in sheets.items():
                        jh_col = None
                        factory_col = None
                        for i, h in enumerate(headers):
                            hs = str(h).strip() if h else ""
                            if hs in ("交货", "交货号"):
                                jh_col = i
                            if hs == "工厂":
                                factory_col = i
                        if jh_col is not None:
                            for row in data_rows:
                                jh_val = str(row[jh_col]).strip() if jh_col < len(row) and row[jh_col] else ""
                                if jh_val and jh_val.isdigit():
                                    jh_to_factory[jh_val] = target_factory
                    found = True
                except Exception:
                    pass
        if not found:
            log(f"警告: 未找到附件 {att_name} 的临时文件")

    if not jh_to_factory:
        return 0

    log(f"工厂标识: {len(jh_to_factory)} 个交货号需覆盖工厂值")

    # 步骤2：打开输出文件，按交货号覆盖工厂字段
    # 覆盖所有含工厂列的sheet：全量数据、筛选数据、交货汇总、交货汇总_文本日期、工厂交货透视
    wb = openpyxl.load_workbook(output_path)
    modified_count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row = [cell.value for cell in ws[1]]

        factory_col = None
        jh_col = None
        for i, h in enumerate(header_row):
            if h and str(h).strip() == "工厂":
                factory_col = i
            if h and str(h).strip() in ("交货", "交货号"):
                jh_col = i

        if factory_col is None or jh_col is None:
            continue

        for row_idx in range(2, ws.max_row + 1):
            jh_val = ws.cell(row=row_idx, column=jh_col + 1).value
            if jh_val is None:
                continue
            jh_str = str(jh_val).strip()
            if jh_str in jh_to_factory:
                ws.cell(row=row_idx, column=factory_col + 1).value = jh_to_factory[jh_str]
                modified_count += 1

    wb.save(output_path)
    wb.close()
    return modified_count


def read_all_sheets_for_override(filepath: str):
    """读取Excel文件的所有sheet，返回 {sheet_name: (headers, data_rows)}
    复用 merger 的逻辑但独立调用，避免循环依赖。
    """
    import sys as _sys
    _sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from merger import read_all_sheets
    return read_all_sheets(filepath)

def matches_keywords(subject: str, keywords: List[str]) -> bool:
    if not keywords:
        return True
    return any(k and k in subject for k in keywords)


def is_excel_attachment(filename: str) -> bool:
    return filename.lower().endswith((".xlsx", ".xls", ".csv", ".tsv"))


def decode_subject(raw) -> str:
    if not raw:
        return ""
    parts = decode_header(raw)
    out = []
    for text, enc in parts:
        if isinstance(text, bytes):
            out.append(text.decode(enc or "utf-8", errors="replace"))
        else:
            out.append(text)
    return "".join(out)


def load_processed_uids(path: str) -> set:
    if not os.path.exists(path):
        return set()
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return set(data)


def save_processed_uids(path: str, uids: set) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(sorted(uids), f)


def filter_new_uids(items: list, processed: set) -> list:
    # items 是 [(folder, uid_bytes), ...]，processed 是 {"folder::uid", ...}
    return [(f, u) for f, u in items if f"{f}::{u.decode()}" not in processed]


# ---------- IMAP 交互 ----------

EXCLUDED_FOLDERS = {"草稿箱", "已发送", "已删除", "垃圾邮件", "病毒邮件", "广告邮件", "订阅邮件", "废弃"}


def utf7_decode(s: str) -> str:
    result = []
    i = 0
    while i < len(s):
        if s[i] == '&':
            end = s.find('-', i)
            if end == -1:
                result.append(s[i:])
                break
            if end == i + 1:
                result.append('&')
            else:
                b64 = s[i + 1:end].replace(',', '/')
                try:
                    raw = base64.b64decode(b64 + '=' * (-len(b64) % 4))
                    result.append(raw.decode('utf-16-be'))
                except Exception:
                    result.append(s[i:end + 1])
            i = end + 1
        else:
            result.append(s[i])
            i += 1
    return ''.join(result)


def list_mail_folders(imap: imaplib.IMAP4_SSL) -> list:
    """返回 [(folder_raw, folder_decoded), ...]，排除系统文件夹"""
    typ, folders = imap.list()
    result = []
    for f in folders:
        raw = f.decode('utf-8', errors='replace')
        m = re.search(r'"([^"]*)"\s*$', raw)
        folder = m.group(1) if m else raw
        decoded = utf7_decode(folder)
        if decoded in EXCLUDED_FOLDERS:
            continue
        result.append((folder, decoded))
    return result


def connect_imap(host: str, email_addr: str, auth_code: str) -> imaplib.IMAP4_SSL:
    imap = imaplib.IMAP4_SSL(host, 993)
    imap.login(email_addr, auth_code)
    # 网易邮箱要求客户端先发送非空 ID（RFC 2971），否则 SELECT 被拒
    imaplib.Commands["ID"] = ("AUTH",)
    imap._simple_command("ID", '("name" "ExcelMerger" "version" "1.0.0" "vendor" "ExcelMerger")')
    return imap


def search_mails(imap: imaplib.IMAP4_SSL, since_date: datetime.date, before_date: datetime.date = None) -> list:
    """搜索所有文件夹（排除系统文件夹），返回 [(folder, uid), ...]"""
    since_str = since_date.strftime("%d-%b-%Y")
    if before_date:
        before_str = before_date.strftime("%d-%b-%Y")
        criteria = f"(SINCE {since_str} BEFORE {before_str})"
    else:
        criteria = f"(SINCE {since_str})"
    results = []
    for folder, _decoded in list_mail_folders(imap):
        try:
            typ, _ = imap.select(f'"{folder}"', readonly=True)
            if typ != "OK":
                continue
            typ, data = imap.uid("search", None, criteria)
            if typ == "OK" and data and data[0]:
                for uid in data[0].split():
                    results.append((folder, uid))
        except Exception:
            continue
    return results


def download_excel_attachments(imap: imaplib.IMAP4_SSL, folder: str, uid: bytes, dest_dir: str, start_idx: int = 0) -> tuple:
    """下载邮件Excel附件，返回 (saved_files, body_text)。

    saved_files: [(path, filename), ...]
    body_text: 邮件正文纯文本（用于工厂关键字识别）
    """
    typ, _ = imap.select(f'"{folder}"', readonly=True)
    if typ != "OK":
        return [], ""
    typ, msg_data = imap.uid("fetch", uid, "(RFC822)")
    if typ != "OK":
        return [], ""
    raw = msg_data[0][1]
    msg = email.message_from_bytes(raw)
    body_text = extract_mail_body(msg)
    saved = []
    idx = start_idx
    for part in msg.walk():
        filename = part.get_filename()
        if not filename:
            continue
        decoded = decode_subject(filename)
        if not is_excel_attachment(decoded):
            continue
        payload = part.get_payload(decode=True)
        if not payload:
            continue
        safe_name = f"{idx}_{decoded}"
        path = os.path.join(dest_dir, safe_name)
        with open(path, "wb") as f:
            f.write(payload)
        saved.append((path, decoded))
        idx += 1
    return saved, body_text


# ---------- 主流程 ----------

_logs = deque(maxlen=200)  # 内存日志，最近 200 条


def log(msg: str) -> None:
    line = f"[{now_cn().strftime('%H:%M:%S')}] {msg}"
    _logs.append(line)
    print(f"[mail_reader] {line}", flush=True)


def get_logs() -> List[str]:
    return list(_logs)


def load_config(path: str) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


MAX_TASKS = 200


def load_tasks(path: str = None) -> list:
    """加载邮件任务历史（优先从数据库，回退到文件）"""
    if _db:
        try:
            return _db.get_mail_tasks()
        except Exception:
            pass
    # 回退到文件（兼容旧数据）
    tasks_file = os.path.join(_base_dir(), "data", "tasks.json")
    if not os.path.exists(tasks_file):
        return []
    try:
        with open(tasks_file, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return []


def save_tasks(tasks: list, path: str = None) -> None:
    """保存邮件任务历史到数据库"""
    if _db:
        try:
            # tasks 是 [{"time":..., "mails":[...]}, ...]
            # database 只存 mails，需要取最新的
            if tasks:
                _db.save_mail_task(tasks[0].get("mails", []))
            return
        except Exception as e:
            print(f"[mail_reader] 保存任务到数据库失败: {e}")
    # 回退到文件
    tasks_file = os.path.join(_base_dir(), "data", "tasks.json")
    d = os.path.dirname(tasks_file)
    if d:
        os.makedirs(d, exist_ok=True)
    with open(tasks_file, "w", encoding="utf-8") as f:
        json.dump(tasks, f, ensure_ascii=False, indent=2)


def clean_output_files(output_dir: str) -> None:
    if not os.path.isdir(output_dir):
        return
    files = [f for f in os.listdir(output_dir) if f.startswith("邮件合并") and f.endswith(".xlsx")]
    today = now_cn().strftime("%Y%m%d")
    by_date = {}
    for f in files:
        m = re.search(r"(\d{8})", f)
        if not m:
            continue
        by_date.setdefault(m.group(1), []).append(f)
    for date, fs in by_date.items():
        if date == today:
            continue
        fs.sort(key=lambda f: os.path.getmtime(os.path.join(output_dir, f)))
        for f in fs[:-1]:
            try:
                os.remove(os.path.join(output_dir, f))
            except OSError:
                pass


def process_once(cfg: dict, force: bool = False) -> int:
    date_str = (cfg.get("date") or "").strip()
    if date_str:
        try:
            target = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            target = now_cn().date()
    else:
        target = now_cn().date()
    since = target
    before = target + datetime.timedelta(days=1)
    # 使用数据库存储已处理 UID（如果没有数据库，回退到文件）
    uids_file = cfg.get("processed_uids_file")
    if uids_file:
        if not os.path.isabs(uids_file):
            uids_file = os.path.join(_base_dir(), uids_file)
        processed = set() if force else load_processed_uids(uids_file)
    elif _db:
        processed = set() if force else _db.get_processed_uids()
    else:
        uids_file = os.path.join(_base_dir(), "data", "processed_uids.json")
        processed = set() if force else load_processed_uids(uids_file)
    output_dir = cfg.get("output_dir", "output")
    if not os.path.isabs(output_dir):
        output_dir = os.path.join(_base_dir(), output_dir)

    log(f"开始一轮：搜索 {target} 的邮件")
    imap = connect_imap(cfg["imap_host"], cfg["email"], cfg["auth_code"])
    task_mails = []
    try:
        items = search_mails(imap, since, before)
        new_items = filter_new_uids(items, processed)
        keywords = cfg.get("subject_keywords", [])
        log(f"共 {len(items)} 封邮件，其中新邮件 {len(new_items)} 封")
        handled = 0
        all_files = []
        _factory_override_map = {}  # 附件名 → 目标工厂值
        collect_dir = tempfile.TemporaryDirectory()
        collect_path = collect_dir.name
        file_idx = 0
        for folder, uid in new_items:
            try:
                imap.select(f'"{folder}"', readonly=True)
                typ, data = imap.uid("fetch", uid, "(BODY[HEADER.FIELDS (SUBJECT)])")
                subject = ""
                if typ == "OK" and data and data[0]:
                    m = email.message_from_bytes(data[0][1])
                    subject = decode_subject(m.get("Subject", ""))
                files, body_text = download_excel_attachments(imap, folder, uid, collect_path, file_idx)
                if not files:
                    log(f"跳过（无 Excel 附件）: {subject}")
                    continue
                attachment_names = [name for _, name in files]
                file_idx += len(files)
                matched = matches_keywords(subject, keywords)
                if matched:
                    all_files.extend([path for path, _ in files])
                    processed.add(f"{folder}::{uid.decode()}")
                    handled += 1
                    # 工厂标识识别：逐个附件检测工厂关键字
                    # 多附件邮件：只从附件名+主题匹配（正文对多附件太宽泛）
                    # 单附件邮件：可以用正文匹配
                    detected_any = False
                    multi_attachment = len(files) > 1
                    for _, att_name in files:
                        # 多附件时不用正文，单附件时用正文
                        body_for_detect = "" if multi_attachment else body_text
                        factory_target = detect_factory_override(subject, [att_name], body_for_detect)
                        if factory_target:
                            _factory_override_map[att_name] = factory_target
                            detected_any = True
                    if detected_any:
                        overrides = ", ".join(f"{k}→{v}" for k, v in _factory_override_map.items() if k in attachment_names)
                        log(f"匹配: {subject} → {len(files)} 个附件 [{overrides}]")
                    else:
                        log(f"匹配: {subject} → {len(files)} 个附件")
                else:
                    log(f"跳过（主题不匹配，含附件）: {subject}")
                task_mails.append({
                    "subject": subject,
                    "attachments": attachment_names,
                    "processed": matched,
                })
            except Exception as e:
                log(f"处理失败: {uid.decode()} - {e}")
        if all_files:
            result = merge_files(
                file_paths=all_files,
                selected_sheets=None,
                provinces=cfg.get("provinces", []),
                rule_id=cfg.get("rule_id"),
                output_dir=output_dir,
                output_prefix=cfg.get("output_prefix", "邮件合并"),
                date_str=target.strftime("%Y-%m-%d"),
            )
            log(f"合并完成: 全量 {result['stats']['total_merged_rows']} 行，筛选 {result['stats']['filtered_rows']} 行")
            # 工厂字段后处理：根据邮件识别到的工厂关键字覆盖工厂值
            if _factory_override_map:
                modified = apply_factory_override(result["output_path"], _factory_override_map, collect_path)
                if modified:
                    log(f"工厂标识覆盖: 修改 {modified} 行")
        else:
            log("无匹配附件，跳过合并")
        if not force:
            if uids_file:
                save_processed_uids(uids_file, processed)
            elif _db:
                _db.add_processed_uids(processed)
        if task_mails:
            task = {
                "time": now_cn().strftime("%Y-%m-%d %H:%M:%S"),
                "mails": task_mails,
            }
            tasks = load_tasks()
            tasks.insert(0, task)
            save_tasks(tasks[:MAX_TASKS])
        clean_output_files(output_dir)
        log(f"本轮完成，处理 {handled} 封邮件")
        return handled
    finally:
        imap.logout()


def main():
    """主函数：从数据库加载配置并启动"""
    if _db:
        cfg = _db.get_mail_config()
    else:
        cfg_path = os.environ.get("MAIL_CONFIG", "mail_config.json")
        cfg = load_config(cfg_path)
    interval = int(cfg.get("poll_interval_seconds", 3600))
    print(f"[mail_reader] 启动，轮询间隔 {interval}s")
    start_background(cfg)
    try:
        while True:
            time.sleep(3600)
    except KeyboardInterrupt:
        stop_background()


# ---------- 后台线程管理 ----------

_stop_event = threading.Event()
_worker_thread = None


def _run_loop(cfg: dict):
    interval = int(cfg.get("poll_interval_seconds", 3600))
    while not _stop_event.is_set():
        try:
            n = process_once(cfg)
            log(f"本轮处理 {n} 封邮件，等待 {interval}s 后进行下一轮")
        except Exception as e:
            log(f"本轮异常: {e}")
        _stop_event.wait(interval)


def start_background(cfg: dict) -> bool:
    global _worker_thread
    if _worker_thread and _worker_thread.is_alive():
        return False
    _stop_event.clear()
    _worker_thread = threading.Thread(target=_run_loop, args=(cfg,), daemon=True)
    _worker_thread.start()
    return True


def stop_background() -> bool:
    global _worker_thread
    if not _worker_thread or not _worker_thread.is_alive():
        return False
    _stop_event.set()
    _worker_thread.join(timeout=5)
    _worker_thread = None
    return True


def is_running() -> bool:
    return _worker_thread is not None and _worker_thread.is_alive()


if __name__ == "__main__":
    main()
