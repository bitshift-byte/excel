"""合并核心逻辑（从 app.py 抽出，供 Web 与邮件读取器共用）"""
import os
import sys
import re
import json
import hashlib
import datetime
from datetime import timezone, timedelta
from collections import OrderedDict
from typing import List, Dict, Tuple, Optional
from copy import copy

import openpyxl
import xlrd

PREVIEW_MAX_ROWS = 200
SAMPLE_ROWS = 10


def _base_dir() -> str:
    """可写数据目录：项目根目录"""
    return os.path.dirname(os.path.abspath(__file__))


def _resource_path(relative: str) -> str:
    """只读资源路径：项目根目录"""
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), relative)


RULES_FILE = os.path.join(_base_dir(), "rules.json")

# ===================== 规则管理 =====================

# 内置默认规则 ID
BUILTIN_RULE_ID = "_builtin_default"

# 内置默认规则：联合利华标准34列 + 列名变体映射
BUILTIN_RULE = {
    "id": BUILTIN_RULE_ID,
    "name": "联合利华标准34列（内置）",
    "builtin": True,
    "standard_headers": [
        {"name": "交货", "source_columns": ["交货", "交货号"]},
        {"name": "DlvTy", "source_columns": ["DlvTy", "交货类型"]},
        {"name": "项目", "source_columns": ["项目", "    项目"]},
        {"name": "物料", "source_columns": ["物料", "物料号"]},
        {"name": "描述", "source_columns": ["描述", "物料描述"]},
        {"name": "存储位置", "source_columns": ["存储位置", "位置"]},
        {"name": "销售凭证", "source_columns": ["销售凭证", "销售订单"]},
        {"name": "运达方", "source_columns": ["运达方", "送达方"]},
        {"name": "运达方的名字", "source_columns": ["运达方的名字", "运达方名称"]},
        {"name": "送达方地点", "source_columns": ["送达方地点", "城市"]},
        {"name": "名称 3", "source_columns": ["名称 3", "名称3"]},
        {
            "name": "工厂",
            "source_columns": ["工厂", "Plant"],
            "value_mappings": [
                {"source_file_contains": "分销下单量", "source_value": "8136", "target_value": "701"},
                {"source_file_contains": "分销下单量", "source_value": "8137", "target_value": "701"},
                {"source_file_contains": "分销报表", "source_value": "8136", "target_value": "901"},
                {"source_file_contains": "分销报表", "source_value": "8137", "target_value": "901"},
                {"source_file_contains": "分销报表", "source_value": "8205", "target_value": "901"},
                {"source_file_contains": "跑单明细", "source_value": "8136", "target_value": "801"},
                {"source_file_contains": "跑单明细", "source_value": "8137", "target_value": "801"},
                {"source_file_contains": "跑单明细", "source_value": "8205", "target_value": "801"},
            ],
        },
        {"name": "路线", "source_columns": ["路线", "Route"]},
        {"name": "OPS", "source_columns": ["OPS", "全部拣配状态"]},
        {"name": "WhN", "source_columns": ["WhN", "仓库号"]},
        {"name": "批次", "source_columns": ["批次", "Batch"]},
        {"name": "仓位", "source_columns": ["仓位"]},
        {"name": "GM", "source_columns": ["GM", "GS", "货物移动状态"]},
        {"name": "销售组织", "source_columns": ["销售组织", "SOrg.", "SOrg"]},
        {"name": "售达方", "source_columns": ["售达方", "售达方代码"]},
        {"name": "售达方的名字", "source_columns": ["售达方的名字", "售达方名称"]},
        {
            "name": "街道",
            "source_columns": ["街道", "街道地址"],
        },
        {"name": "街道2", "source_columns": ["街道2", "街道 2"]},
        {"name": "街道 3", "source_columns": ["街道 3", "街道3"]},
        {"name": "交货量", "source_columns": ["交货量", "交货数量", "    交货量"]},
        {"name": "SU", "source_columns": ["SU", "销售单位"]},
        {"name": "数量(库存单位)", "source_columns": ["数量(库存单位)", "库存数量"]},
        {"name": "计", "source_columns": ["计", "计数", "基本计量单位"]},
        {"name": "总重量", "source_columns": ["总重量", "         总重量"]},
        {"name": "WUn", "source_columns": ["WUn", "重量单位"]},
        {"name": "业务量", "source_columns": ["业务量", "          业务量"]},
        {"name": "VUn", "source_columns": ["VUn", "体积单位"]},
        {"name": "交货日期", "source_columns": ["交货日期", "交货日期(从/到)"]},
        {"name": "发货日期", "source_columns": ["发货日期", "实际发货日", "实际货物移动日期"]},
    ],
}


def load_rules() -> list:
    """从 rules.json 读取规则列表，始终在列表首位插入内置默认规则"""
    user_rules = []
    if os.path.exists(RULES_FILE):
        try:
            with open(RULES_FILE, "r", encoding="utf-8") as f:
                user_rules = json.load(f)
        except (json.JSONDecodeError, IOError):
            user_rules = []
    # 内置规则始终在最前面
    return [BUILTIN_RULE] + user_rules


def save_rules(rules: list):
    """将规则列表写入 rules.json（自动过滤内置规则）"""
    user_rules = [r for r in rules if not r.get("builtin")]
    with open(RULES_FILE, "w", encoding="utf-8") as f:
        json.dump(user_rules, f, ensure_ascii=False, indent=2)


def normalize_str(s: str) -> str:
    """标准化字符串用于模糊匹配：去首尾空格 + 转小写 + 去内部空格/下划线"""
    if not s:
        return ""
    return s.strip().lower().replace(" ", "").replace("_", "")


def match_columns_to_rule(headers: list, rule: dict) -> list:
    """
    将表头列表与规则中的标准表头匹配。
    返回映射列表: [(原始列名, 标准表头名), ...]，长度等于 headers，未匹配的为 None。
    支持重复列名：同一列名出现多次时可分别映射到不同标准列。

    匹配策略：
    1. 先精确匹配（normalize 后相等）
    2. 再包含匹配（normalize 后 source_column 是 header 的子串，或反之）
    3. 一个表头只匹配第一个命中的标准表头
    4. 多个表头匹配同一标准表头时取第一个，其余尝试下一个标准表头
    """
    std_headers = rule.get("standard_headers", [])
    # 返回 list of (header_text, std_name or None)
    result = []
    used_targets = set()  # 已被占用的标准表头名

    for header in headers:
        hs = str(header) if header else ""
        if not hs:
            result.append((hs, None))
            continue
        hs_norm = normalize_str(hs)
        matched = None

        # Pass 1: 精确匹配
        for sh in std_headers:
            target = sh.get("name", "")
            if not target or target in used_targets:
                continue
            for sc in sh.get("source_columns", []):
                if normalize_str(sc) == hs_norm:
                    matched = target
                    break
            if matched:
                break

        # Pass 2: 包含匹配（要求较长的匹配长度，避免短词误匹配）
        if not matched:
            for sh in std_headers:
                target = sh.get("name", "")
                if not target or target in used_targets:
                    continue
                for sc in sh.get("source_columns", []):
                    sc_norm = normalize_str(sc)
                    if not sc_norm:
                        continue
                    # 包含匹配：要求匹配长度至少为较长一方的 50%
                    max_len = max(len(sc_norm), len(hs_norm))
                    min_len = min(len(sc_norm), len(hs_norm))
                    if sc_norm in hs_norm or hs_norm in sc_norm:
                        if min_len >= max_len * 0.5:
                            matched = target
                            break
                if matched:
                    break

        if matched:
            used_targets.add(matched)
            result.append((hs, matched))
        else:
            result.append((hs, None))

    return result


def apply_value_mappings(row_dict: dict, std_name: str, value_mappings: list, filename: str) -> bool:
    """对某行的指定标准列应用值映射规则。如果有变更返回 True 并直接修改 row_dict。

    支持两种映射类型：
    1. 源文件名 + 源值 → 目标值：
       {"source_file_contains": "分销下单量", "source_value": "8136", "target_value": "701"}
    2. 条件跨列映射（当某列等于某值时，用另一列的值替换）：
       {"when_column": "工厂", "equals": "901", "use_column": "送达方地点"}
    """
    if not value_mappings:
        return False
    current_val = row_dict.get(std_name, "")
    current_str = str(current_val).strip() if current_val is not None else ""

    for vm in value_mappings:
        # Type 1: source file + source value -> target value
        if "source_file_contains" in vm and "source_value" in vm:
            file_keyword = vm.get("source_file_contains", "")
            src_val = str(vm.get("source_value", "")).strip()
            if file_keyword in filename and current_str == src_val:
                row_dict[std_name] = vm.get("target_value", "")
                return True
        # Type 2: conditional cross-column mapping
        elif "when_column" in vm and "equals" in vm:
            when_col = vm.get("when_column", "")
            when_val = str(row_dict.get(when_col, "")).strip() if row_dict.get(when_col) is not None else ""
            equals_val = str(vm.get("equals", "")).strip()
            if when_val == equals_val:
                if vm.get("use_column"):
                    row_dict[std_name] = row_dict.get(vm["use_column"], "")
                    return True
                elif vm.get("target_value") is not None:
                    row_dict[std_name] = vm.get("target_value", "")
                    return True
    return False


# ===================== 行政区划数据 =====================

REGIONS_FILE = _resource_path("china_regions.json")


def build_region_keywords() -> Dict[str, list]:
    with open(REGIONS_FILE, "r", encoding="utf-8") as f:
        raw = json.load(f)

    regions = {}
    for prov, cities in raw.items():
        keywords = set()
        keywords.add(prov)
        short_prov = prov
        for suffix in ["省", "自治区", "壮族", "回族", "维吾尔"]:
            short_prov = short_prov.replace(suffix, "")
        if short_prov:
            keywords.add(short_prov)

        for city, counties in cities.items():
            keywords.add(city)
            short_city = city
            for suffix in ["市", "自治州", "地区", "盟"]:
                short_city = short_city.replace(suffix, "")
            if short_city:
                keywords.add(short_city)

            for county in counties:
                keywords.add(county)
                for suffix in ["区", "县", "市", "旗", "自治县", "自治区"]:
                    if county.endswith(suffix) and len(county) > len(suffix) + 1:
                        keywords.add(county[: -len(suffix)])

        keywords = sorted([k for k in keywords if len(k) >= 2])
        regions[prov] = keywords

    return regions


REGION_KEYWORDS = build_region_keywords()


def get_province_list() -> list:
    return [
        {"name": prov, "keyword_count": len(kws)}
        for prov, kws in sorted(REGION_KEYWORDS.items())
    ]


def match_province(value, provinces: list) -> bool:
    """检查某个值（城市名/街道地址）是否匹配选中省份"""
    if not value or not provinces:
        return False
    s = str(value).strip()
    if not s:
        return False
    for prov in provinces:
        kws = REGION_KEYWORDS.get(prov, [])
        for kw in kws:
            if s.startswith(kw):
                return True
    return False


def match_row_province(row: dict, provinces: list) -> bool:
    """检查一行数据是否匹配选中省份：优先看'送达方地点'，再看'街道'"""
    if not provinces:
        return True  # 未选省份时全部通过
    # 优先检查"送达方地点"（城市级别，用 startswith 精确匹配）
    for col_name in ["送达方地点", "城市"]:
        val = row.get(col_name)
        if val and match_province(val, provinces):
            return True
    # 再检查"街道"地址（地址通常以省名开头，用 startswith 匹配省份名）
    for col_name in ["街道", "名称 3", "街道 3"]:
        val = row.get(col_name)
        if val:
            s = str(val).strip()
            for prov in provinces:
                kws = REGION_KEYWORDS.get(prov, [])
                for kw in kws:
                    if s.startswith(kw):
                        return True
    return False



# 中国时区 UTC+8
_CN_TZ = timezone(timedelta(hours=8))

def now_cn() -> datetime.datetime:
    """返回中国时区(UTC+8)的当前时间"""
    return datetime.datetime.now(_CN_TZ)

def fromtimestamp_cn(ts: float) -> datetime.datetime:
    """将时间戳转为中国时区(UTC+8)的 datetime"""
    return datetime.datetime.fromtimestamp(ts, tz=_CN_TZ)



def _excel_date(val):
    """将 datetime 值转为 date-only，避免 Excel 显示时分秒"""
    if isinstance(val, datetime.datetime):
        return datetime.date(val.year, val.month, val.day)
    return val


def normalize_date_formats(wb):
    """统一设置 workbook 所有 sheet 中日期单元格的显示格式为斜杠 yyyy/m/d，
    同时加宽含日期但列宽不足的列（yyyy/m/d 需要 ≥12 字符宽度）。
    在 wb.save() 之前调用。"""
    _DATE_FMT = "yyyy/m/d"
    _MIN_DATE_COL_WIDTH = 12
    for ws in wb.worksheets:
        date_cols = set()
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (datetime.datetime, datetime.date)):
                    cell.number_format = _DATE_FMT
                    date_cols.add(cell.column_letter)
        for col_letter in date_cols:
            dim = ws.column_dimensions.get(col_letter)
            current_width = dim.width if dim and dim.width else 8.89
            if current_width < _MIN_DATE_COL_WIDTH:
                ws.column_dimensions[col_letter].width = _MIN_DATE_COL_WIDTH


def serialize_cell(val):
    if val is None:
        return ""
    if isinstance(val, datetime.datetime):
        return val.strftime("%Y/%m/%d")
    if isinstance(val, datetime.date):
        return val.strftime("%Y/%m/%d")
    return str(val)


def _to_number(val):
    """尝试将值转为 float，失败返回 None"""
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _format_date_text(val) -> str:
    """将日期值转为 'YYYY.MM.DD' 文本格式"""
    if val is None or val == "":
        return ""
    if isinstance(val, datetime.datetime):
        return val.strftime("%Y/%m/%d")
    if isinstance(val, datetime.date):
        return val.strftime("%Y/%m/%d")
    s = str(val).strip()
    # 尝试解析常见的日期格式
    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"]:
        try:
            dt = datetime.datetime.strptime(s, fmt)
            return dt.strftime("%Y/%m/%d")
        except ValueError:
            pass
    return s


def _try_parse_date(val):
    """尝试将值转为 datetime，失败则返回 None"""
    if val is None or (isinstance(val, str) and not val.strip()):
        return None
    if isinstance(val, datetime.datetime):
        return val
    if isinstance(val, datetime.date):
        return datetime.datetime(val.year, val.month, val.day)
    s = str(val).strip()
    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"]:
        try:
            return datetime.datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def build_pivot_by_delivery(filtered_rows: list, logistics_map: Optional[Dict] = None, so_map: Optional[Dict] = None, kuacang_map: Optional[Dict] = None) -> Tuple[list, list, dict]:
    """Sheet4/Sheet5: 按交货号汇总透视表（从 Sheet3 筛选数据构建）
    列: 交货, 销售凭证, 运达方, 运达方的名字, 送达方地点, 工厂, 街道, 发货日期, 交货日期, 求和项:交货量, 求和项:总重量, 求和项:业务量, B_ADDRESS1, 备注, 1, 2

    Sheet4 规则（模拟 Excel 透视表行为）：
    - 数值列：SUM 求和
    - 日期为 datetime → Sheet4 保持 datetime，街道保持原值
    - 日期为文本格式(str) → Sheet4 中日期=None，街道=None（Excel 透视表丢弃文本日期行字段）
    - B_ADDRESS1/备注：先查 logistics_map（已发运/未发运，key=销售凭证），未命中的再查 so_map（SO文件，key=交货号），再未命中查 kuacang_map（跨仓订单，key=交货号）
    - 列1/列2（O/P 子合计）：该交货下"奥妙+洗衣粉/皂粉"类别的总重量/业务量子合计

    Sheet5 规则：Sheet4 的副本，但 None 的日期/街道用 Sheet3 原始文本值回填
    """
    pivot_map = OrderedDict()
    # 保存 Sheet3 中的原始文本值（用于 Sheet5 回填）
    sheet3_orig = {}  # delivery -> {"发货日期": orig_val, "交货日期": orig_val, "街道": orig_val}
    # 计算"奥妙+洗衣粉/皂粉"子合计（用于 Sheet4 的 O/P 列和 Sheet5）
    omo_subtotals = {}  # delivery -> {"zzl": 0.0, "ywl": 0.0}

    for row in filtered_rows:
        delivery = str(row.get("交货", "")).strip()
        if not delivery:
            continue
        fa_val = row.get("发货日期", "")
        jr_val = row.get("交货日期", "")
        # 过滤掉日期为空的行（与标准答案对齐，避免无日期的交货号进入汇总）
        if not fa_val or not str(fa_val).strip():
            continue
        if not jr_val or not str(jr_val).strip():
            continue
        street_val = row.get("街道", "")

        # 检查是否为"奥妙+洗衣粉/皂粉"类别
        desc = str(row.get("描述", ""))
        is_omo = "奥妙" in desc and ("洗衣粉" in desc or "皂粉" in desc)
        if is_omo:
            if delivery not in omo_subtotals:
                omo_subtotals[delivery] = {"zzl": 0.0, "ywl": 0.0}
            zzl_v = _to_number(row.get("总重量"))
            ywl_v = _to_number(row.get("业务量"))
            if zzl_v:
                omo_subtotals[delivery]["zzl"] += zzl_v
            if ywl_v:
                omo_subtotals[delivery]["ywl"] += ywl_v

        # 收集原始值（用于 Sheet5 回填）
        if delivery not in sheet3_orig:
            sheet3_orig[delivery] = {}
        if fa_val and str(fa_val).strip() and not sheet3_orig[delivery].get("发货日期"):
            sheet3_orig[delivery]["发货日期"] = fa_val
        if jr_val and str(jr_val).strip() and not sheet3_orig[delivery].get("交货日期"):
            sheet3_orig[delivery]["交货日期"] = jr_val
        if street_val and str(street_val).strip() and not sheet3_orig[delivery].get("街道"):
            sheet3_orig[delivery]["街道"] = street_val

        if delivery not in pivot_map:
            # 判断日期是 datetime 还是文本
            fa_is_dt = isinstance(fa_val, datetime.datetime)
            jr_is_dt = isinstance(jr_val, datetime.datetime)
            pivot_map[delivery] = {
                "交货": delivery,
                "销售凭证": row.get("销售凭证", ""),
                "运达方": row.get("运达方", ""),
                "运达方的名字": row.get("运达方的名字", ""),
                "送达方地点": row.get("送达方地点", ""),
                "工厂": row.get("工厂", ""),
                # 街道始终保留原值（与标准答案行为一致）
                "街道": street_val,
                "发货日期": datetime.date(fa_val.year, fa_val.month, fa_val.day) if fa_is_dt else None,
                "交货日期": datetime.date(jr_val.year, jr_val.month, jr_val.day) if jr_is_dt else None,
                "_交货量": 0.0,
                "_总重量": 0.0,
                "_业务量": 0.0,
                "_count": 0,
            }
        entry = pivot_map[delivery]
        entry["_交货量"] += _to_number(row.get("交货量")) or 0
        entry["_总重量"] += _to_number(row.get("总重量")) or 0
        entry["_业务量"] += _to_number(row.get("业务量")) or 0
        entry["_count"] += 1

    pivot_headers = [
        "发货日期", "交货日期", "送达方地点", "运达方", "销售凭证",
        "交货", "运达方的名字", "街道",
        "求和项:交货量", "求和项:总重量", "求和项:业务量", "工厂",
        "B_ADDRESS1", "备注", "1", "2",
    ]
    result = []
    total_jhl = 0.0
    total_zzl = 0.0
    total_ywl = 0.0
    for entry in pivot_map.values():
        jhl = round(entry["_交货量"], 3) if entry["_交货量"] else 0
        zzl = round(entry["_总重量"], 3) if entry["_总重量"] else 0
        ywl = round(entry["_业务量"], 3) if entry["_业务量"] else 0
        total_jhl += jhl
        total_zzl += zzl
        total_ywl += ywl
        # 通过 logistics_map LEFT JOIN 取 B_ADDRESS1/备注
        addr1 = ""
        remark = ""
        if logistics_map:
            xp = str(entry["销售凭证"]).strip() if entry["销售凭证"] else ""
            logi = logistics_map.get(xp)
            if logi:
                addr1 = logi.get("B_ADDRESS1", "") or ""
                remark = logi.get("备注", "") or ""
        # 2. 如果 logistics_map 未命中，用交货号查 SO map
        # SO 文件的"客户订单号"字段值 == 本表的"交货"号
        if so_map:
            delivery_no = str(entry["交货"]).strip() if entry["交货"] else ""
            so_data = so_map.get(delivery_no)
            if so_data:
                if not addr1:
                    addr1 = so_data.get("B_ADDRESS1", "") or ""
                if not remark:
                    remark = so_data.get("备注", "") or ""
        # 3. 如果仍未命中，用交货号查 跨仓订单 map（kuacang_map）
        if kuacang_map:
            delivery_no = str(entry["交货"]).strip() if entry["交货"] else ""
            kc_data = kuacang_map.get(delivery_no)
            if kc_data:
                if not addr1:
                    addr1 = kc_data.get("B_ADDRESS1", "") or ""
                if not remark:
                    remark = kc_data.get("备注", "") or ""
        # O/P 子合计：奥妙+洗衣粉/皂粉的总重量/业务量
        omo = omo_subtotals.get(entry["交货"])
        o_val = round(omo["zzl"], 3) if omo else None
        p_val = round(omo["ywl"], 3) if omo else None
        result.append([
            entry["发货日期"],
            entry["交货日期"],
            entry["送达方地点"],
            entry["运达方"],
            entry["销售凭证"],
            entry["交货"],
            entry["运达方的名字"],
            entry["街道"],
            jhl,
            zzl,
            ywl,
            entry["工厂"],
            remark,
            addr1,
            o_val,
            p_val,
        ])
    # 添加 "总计" 行
    result.append(["总计", None, None, None, None, None, None, None, round(total_jhl, 3), round(total_zzl, 3), round(total_ywl, 3), None, None, None, None, None])
    return pivot_headers, result, sheet3_orig



def build_data_pivot(p4_headers: list, p4_data: list) -> Tuple[list, list]:
    """构建"数据透析" sheet 数据。

    数据透析是交货汇总的重排+精简版本：
    - 12列（丢弃 B_ADDRESS1, 备注, 1, 2）
    - 列顺序: 交货, 销售凭证, 运达方, 运达方的名字, 送达方地点, 工厂, 街道,
              发货日期, 交货日期, 求和项:交货量, 求和项:总重量, 求和项:业务量
    - 按交货号升序排列
    - 数据行后追加 (空白) 行和 总计 行

    Args:
        p4_headers: 交货汇总的 16 列表头
        p4_data: 交货汇总的数据行列表（最后一行为 总计 行）

    Returns:
        (dp_headers, dp_data): 数据透析的 12 列表头和数据行
    """
    # 交货汇总 16-col → 数据透析 12-col 的列映射
    # 交货汇总 index: 交货[5], 销售凭证[4], 运达方[3], 运达方的名字[6],
    #                 送达方地点[2], 工厂[11], 街道[7], 发货日期[0], 交货日期[1],
    #                 求和项:交货量[8], 求和项:总重量[9], 求和项:业务量[10]
    COL_MAP = [5, 4, 3, 6, 2, 11, 7, 0, 1, 8, 9, 10]

    dp_headers = [
        "交货", "销售凭证", "运达方", "运达方的名字", "送达方地点",
        "工厂", "街道", "发货日期", "交货日期",
        "求和项:交货量", "求和项:总重量", "求和项:业务量",
    ]

    # 分离数据行和总计行
    data_rows = []
    total_row = None
    for row in p4_data:
        if row and str(row[0]).strip() == "总计":
            total_row = row
        else:
            # 重排列并截取 12 列
            new_row = [row[src_idx] if src_idx < len(row) else None for src_idx in COL_MAP]
            data_rows.append(new_row)

    # 按交货号（index 0）升序排序
    data_rows.sort(key=lambda r: str(r[0]) if r[0] is not None else "")

    # 从总计行提取合计值（交货汇总 index 8/9/10）
    if total_row:
        total_jhl = total_row[8] if len(total_row) > 8 and total_row[8] is not None else 0
        total_zzl = total_row[9] if len(total_row) > 9 and total_row[9] is not None else 0
        total_ywl = total_row[10] if len(total_row) > 10 and total_row[10] is not None else 0
    else:
        total_jhl = 0
        total_zzl = 0
        total_ywl = 0

    # 追加 (空白) 行
    data_rows.append(["(空白)"] * 9 + [None, None, None])

    # 追加 总计 行
    data_rows.append(["总计", None, None, None, None, None, None, None, None,
                      total_jhl, total_zzl, total_ywl])

    return dp_headers, data_rows

def read_logistics_map(files_data: Dict) -> Dict[str, Dict]:
    """从已发运/未发运 sheet 构建 销售凭证 → {B_ADDRESS1, 备注} 映射。
    
    遍历所有输入文件中名为"已发运"或"未发运"的 sheet，
    按"销售凭证"列去重（LEFT JOIN 语义：已发运优先）。
    """
    logistics_map = {}
    for fname, sheets in files_data.items():
        for sname, (headers, data_rows) in sheets.items():
            if sname not in ("已发运", "未发运"):
                continue
            h_map = {}
            for i, h in enumerate(headers):
                h_map[str(h).strip() if h else ""] = i
            ci_xp = h_map.get("销售凭证")
            ci_addr = h_map.get("B_ADDRESS1")
            ci_remark = h_map.get("备注")
            if ci_xp is None:
                continue
            is_sent = (sname == "已发运")
            for row in data_rows:
                xp = str(row[ci_xp]).strip() if ci_xp < len(row) and row[ci_xp] is not None else ""
                if not xp:
                    continue
                # 已发运优先：如果已存在则不覆盖
                if xp in logistics_map and logistics_map[xp].get("_from_sent"):
                    continue
                addr = row[ci_addr] if ci_addr is not None and ci_addr < len(row) else None
                remark = row[ci_remark] if ci_remark is not None and ci_remark < len(row) else None
                logistics_map[xp] = {
                    "B_ADDRESS1": addr if addr is not None else "",
                    "备注": remark if remark is not None else "",
                    "_from_sent": is_sent,
                }
    return logistics_map


def read_so_map(files_data: Dict) -> Dict[str, Dict]:
    """从 SO 文件构建 客户订单号(=交货号) → {B_ADDRESS1, 备注} 映射。

    SO 文件特征：
    - sheet 名为 Sheet1/Sheet2 等（非"已发运"/"未发运"）
    - 表头同时含"客户订单号"、"NOTES2"、"B_ADDRESS1"
    - JOIN key: "客户订单号"（值 == 产出物的"交货"号）
    - NOTES2 → 备注, B_ADDRESS1 → B_ADDRESS1
    """
    so_map = {}
    for fname, sheets in files_data.items():
        for sname, (headers, data_rows) in sheets.items():
            h_map = {}
            for i, h in enumerate(headers):
                h_map[str(h).strip() if h else ""] = i
            ci_order = h_map.get("客户订单号")
            # NOTES2 / NOTE2, B_ADDRESS1 / 地址 是同义列名变体（SO vs OMR）
            ci_notes = h_map.get("NOTES2", h_map.get("NOTE2"))
            ci_addr = h_map.get("B_ADDRESS1", h_map.get("地址"))
            # 必须同时含三列才认定为 SO 文件
            if ci_order is None or ci_notes is None or ci_addr is None:
                continue
            for row in data_rows:
                order_no = str(row[ci_order]).strip() if ci_order < len(row) and row[ci_order] is not None else ""
                if not order_no:
                    continue
                # 有值优先：已存在且有 B_ADDRESS1 值则不覆盖
                if order_no in so_map and so_map[order_no].get("B_ADDRESS1"):
                    continue
                notes_val = row[ci_notes] if ci_notes < len(row) else None
                addr_val = row[ci_addr] if ci_addr < len(row) else None
                so_map[order_no] = {
                    "B_ADDRESS1": addr_val if addr_val is not None else "",
                    "备注": notes_val if notes_val is not None else "",
                }
    return so_map


def read_kuacang_map(files_data: Dict) -> Dict[str, Dict]:
    """从跨仓订单文件构建 OBD(=交货号) → {B_ADDRESS1, 备注} 映射。

    跨仓订单文件特征：
    - sheet 名为 "跨仓结果仓库回传"（允许前后空白）
    - 表头含 "OBD"、"客户订单号"、"备注"、"仓库备注"
    - OBD 值 == 产出物的"交货"号
    - 工单号 → B_ADDRESS1（优先于客户订单号，不拼接）
    - 填写人 → 备注（优先于备注/仓库备注，不拼接）

    优先级：已存在且有 B_ADDRESS1 值的条目不被覆盖 B_ADDRESS1；
    已存在且有备注的条目不被覆盖备注；空字段可被后续行补全。
    """
    kuacang_map: Dict[str, Dict] = {}
    for fname, sheets in files_data.items():
        for sname, (headers, data_rows) in sheets.items():
            # 允许 sheet 名前后有空白
            if sname.strip() != "跨仓结果仓库回传":
                continue
            h_map = {}
            for i, h in enumerate(headers):
                key = str(h).strip() if h else ""
                # 保留第一次出现的列索引（"OBD" 在跨仓订单中出现两次，取第一个）
                if key not in h_map:
                    h_map[key] = i
            # 需要同时有 OBD 和 客户订单号 才认定为跨仓订单 sheet
            ci_obd = h_map.get("OBD")
            ci_cust = h_map.get("客户订单号")
            if ci_obd is None or ci_cust is None:
                continue
            ci_remark = h_map.get("备注")
            ci_wh_remark = h_map.get("仓库备注")
            ci_gdh = h_map.get("工单号")
            ci_txr = h_map.get("填写人")

            for row in data_rows:
                obd_raw = row[ci_obd] if ci_obd < len(row) and row[ci_obd] is not None else ""
                # 规范化 OBD：float → int 字符串（xlrd 读取 .xls 时数字为 float）
                if isinstance(obd_raw, float) and obd_raw.is_integer():
                    obd = str(int(obd_raw))
                else:
                    obd = str(obd_raw).strip() if obd_raw else ""
                # 跳过空/None/0 的 OBD
                if not obd or obd == "0":
                    continue
                cust_po = str(row[ci_cust]).strip() if ci_cust < len(row) and row[ci_cust] is not None else ""
                # B_ADDRESS1: 工单号优先，无则用客户订单号（不拼接）
                gdh_val = ""
                if ci_gdh is not None and ci_gdh < len(row):
                    v = row[ci_gdh]
                    if v is not None and str(v).strip():
                        gdh_val = str(v).strip()
                addr1 = gdh_val if gdh_val else cust_po
                # 备注: 填写人优先，无则用 备注 + 仓库备注（拼接）
                txr_val = ""
                if ci_txr is not None and ci_txr < len(row):
                    v = row[ci_txr]
                    if v is not None and str(v).strip():
                        txr_val = str(v).strip()
                if txr_val:
                    remark = txr_val
                else:
                    parts = []
                    if ci_remark is not None and ci_remark < len(row):
                        v = row[ci_remark]
                        if v is not None and str(v).strip():
                            parts.append(str(v).strip())
                    if ci_wh_remark is not None and ci_wh_remark < len(row):
                        v = row[ci_wh_remark]
                        if v is not None and str(v).strip():
                            parts.append(str(v).strip())
                    remark = " | ".join(parts) if parts else ""
                # 客户订单号和备注都为空，跳过此行（无有用数据）
                if not addr1 and not remark:
                    continue
                # 已存在的条目：补全空字段，不覆盖已有值
                if obd in kuacang_map:
                    existing = kuacang_map[obd]
                    if not existing.get("B_ADDRESS1") and addr1:
                        existing["B_ADDRESS1"] = addr1
                    if not existing.get("备注") and remark:
                        existing["备注"] = remark
                else:
                    kuacang_map[obd] = {
                        "B_ADDRESS1": addr1,
                        "备注": remark,
                    }
    return kuacang_map



def build_omo_detail(filtered_rows: list, std_headers: list) -> Tuple[list, list]:
    """生成"奥妙+洗衣粉/皂粉"明细子集和小计表。
    
    返回 (omo_detail_rows, omo_subtotal_rows):
    - omo_detail_rows: 筛选出的明细行（同标准列，去掉空列），用于"奥妙明细"sheet
    - omo_subtotal_rows: 按交货号汇总的 (交货, 求和项:总重量, 求和项:业务量)，用于"奥妙小计"sheet
    """
    detail_rows = []
    subtotal_map = OrderedDict()
    
    for row in filtered_rows:
        desc = str(row.get("描述", ""))
        if "奥妙" not in desc or ("洗衣粉" not in desc and "皂粉" not in desc):
            continue
        delivery = str(row.get("交货", "")).strip()
        if not delivery:
            continue
        # 明细行
        detail_rows.append([row.get(h, "") for h in std_headers])
        # 小计
        if delivery not in subtotal_map:
            subtotal_map[delivery] = {"zzl": 0.0, "ywl": 0.0}
        zzl_v = _to_number(row.get("总重量"))
        ywl_v = _to_number(row.get("业务量"))
        if zzl_v:
            subtotal_map[delivery]["zzl"] += zzl_v
        if ywl_v:
            subtotal_map[delivery]["ywl"] += ywl_v
    
    subtotal_rows = []
    for delivery, sub in subtotal_map.items():
        subtotal_rows.append([
            delivery,
            round(sub["zzl"], 3) if sub["zzl"] else 0,
            round(sub["ywl"], 3) if sub["ywl"] else 0,
        ])
    
    return detail_rows, subtotal_rows


def build_pivot_by_factory_delivery(all_rows: list) -> Tuple[list, list]:
    """Sheet2: 按工厂+交货号透视
    列: 工厂, 交货, 计数项:物料, 求和项:交货量, 求和项:总重量

    返回结构匹配标准答案的Excel透视表格式:
    - 前2行为空行
    - 第3行: 列C="值"
    - 第4行: 表头
    - 数据行: 工厂名仅在该组首行显示
    - 最后一行: 总计行
    """
    pivot_map = OrderedDict()
    for row in all_rows:
        factory = str(row.get("工厂", "")).strip()
        delivery = str(row.get("交货", "")).strip()
        if not delivery:
            continue
        key = (factory, delivery)
        if key not in pivot_map:
            pivot_map[key] = {
                "工厂": factory,
                "交货": delivery,
                "_交货量": 0.0,
                "_总重量": 0.0,
                "_material_count": 0,
            }
        entry = pivot_map[key]
        material = str(row.get("物料", "")).strip()
        if material:
            entry["_material_count"] += 1
        entry["_交货量"] += _to_number(row.get("交货量")) or 0
        entry["_总重量"] += _to_number(row.get("总重量")) or 0

    pivot_headers = ["工厂", "交货", "计数项:物料", "求和项:交货量", "求和项:总重量"]

    # 按工厂分组排序（工厂名排序，同工厂内按交货号排序）
    sorted_items = sorted(pivot_map.items(), key=lambda x: (x[0][0], x[0][1]))

    result = []
    prev_factory = None
    total_materials = 0
    total_jhl = 0.0
    total_zzl = 0.0
    for (factory, delivery), entry in sorted_items:
        factory_display = factory if factory != prev_factory else ""
        mat_count = entry["_material_count"]
        jhl = round(entry["_交货量"], 3) if entry["_交货量"] else 0
        zzl = round(entry["_总重量"], 3) if entry["_总重量"] else 0
        result.append([
            factory_display,
            delivery,
            mat_count,
            jhl,
            zzl,
        ])
        total_materials += mat_count
        total_jhl += jhl
        total_zzl += zzl
        prev_factory = factory

    # 添加总计行
    result.append([
        "总计",
        None,
        total_materials,
        round(total_jhl, 3) if total_jhl else 0,
        round(total_zzl, 3) if total_zzl else 0,
    ])

    # 构建完整输出: 2空行 + "值"行 + 表头 + 数据 + 总计
    # 透视表格式: 前2行空, 第3行列C="值", 第4行表头
    full_result = []
    # 行0: 空行
    full_result.append([None, None, None, None, None])
    # 行1: 空行
    full_result.append([None, None, None, None, None])
    # 行2: 列C="值" (模拟Excel透视表标记)
    full_result.append([None, None, "值", None, None])
    # 行3: 表头
    full_result.append(pivot_headers[:])
    # 行4+: 数据行 + 总计行
    full_result.extend(result)

    return pivot_headers, full_result


def _normalize_cell(value):
    """规范化单元格值：将整数浮点数转为整数（xlrd 读取 .xls 时数字均为 float），
    将文本格式日期（如 "2026.08.18"）解析为 datetime 对象。"""
    if isinstance(value, float) and value.is_integer():
        return int(value)
    # 尝试解析文本格式日期：YYYY.MM.DD 或 YYYY-MM.DD 等
    if isinstance(value, str):
        s = value.strip()
        if s and re.match(r'^\d{4}[.\-/]\d{1,2}[.\-/]\d{1,2}$', s):
            for sep in ('.', '-', '/'):
                parts = s.split(sep)
                if len(parts) == 3:
                    try:
                        y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
                        return datetime.datetime(y, m, d)
                    except (ValueError, IndexError):
                        pass
    return value


def _parse_sheet_rows(rows: list) -> Tuple[tuple, list]:
    """从行列表中提取表头和有效数据行（不含空行）。

    自动跳过开头的标题行 / 空行：扫描前 10 行，选择非空单元格最多的一行作为表头。
    真正的表头行通常有大量非空列，而标题行只有 1-3 个非空单元格。
    """
    if not rows:
        return None

    def _non_empty_count(row):
        return sum(1 for c in row if c is not None and str(c).strip() != "")

    # 扫描前 10 行，选择非空单元格最多的一行作为表头
    max_scan = min(len(rows), 10)
    header_idx = 0
    best_count = 0
    for i in range(max_scan):
        cnt = _non_empty_count(rows[i])
        if cnt > best_count:
            best_count = cnt
            header_idx = i

    raw_headers = rows[header_idx]
    # 找到第一个和最后一个非空列，剥离前导/尾部空列
    first_non_empty = None
    last_non_empty = -1
    for idx, h in enumerate(raw_headers):
        if h is not None and str(h).strip() != "":
            if first_non_empty is None:
                first_non_empty = idx
            last_non_empty = idx
    if first_non_empty is None or last_non_empty < 0:
        return None
    headers = tuple(_normalize_cell(h) for h in raw_headers[first_non_empty : last_non_empty + 1])
    data_rows = [
        tuple(_normalize_cell(c) for c in r[first_non_empty : last_non_empty + 1])
        for r in rows[header_idx + 1 :]
    ]
    data_rows = [r for r in data_rows if any(c is not None and str(c).strip() != "" for c in r)]
    return headers, data_rows


def _detect_file_type(filepath: str) -> str:
    """通过 magic bytes 检测文件真实类型，而非仅依赖扩展名。
    返回: 'xlsx' | 'xls' | 'tsv' | 'csv' | 'unknown'
    """
    with open(filepath, "rb") as f:
        head = f.read(8)
    # ZIP magic (xlsx / real xls with macro)
    if head[:4] == b"PK\x03\x04":
        return "xlsx"
    # OLE2 compound document (real .xls binary)
    if head[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return "xls"
    # UTF-16 LE BOM
    if head[:2] == b"\xff\xfe":
        return "tsv"
    # UTF-8 BOM or plain text — treat as csv/tsv
    if head[:3] == b"\xef\xbb\xbf" or head[:1] != b"\x00":
        # try to distinguish csv vs tsv by counting tabs vs commas in first 4KB
        return "csv"  # will auto-detect delimiter later
    return "unknown"


def _read_text_table(filepath: str) -> Dict[str, Tuple[tuple, list]]:
    """读取 TSV/CSV 文本文件（可能是 UTF-16 或 UTF-8 编码），返回单 sheet 结果"""
    import csv
    import io

    # 自动检测编码
    with open(filepath, "rb") as f:
        raw = f.read()

    encoding = "utf-8"
    if raw[:2] == b"\xff\xfe":
        encoding = "utf-16-le"
    elif raw[:2] == b"\xfe\xff":
        encoding = "utf-16-be"
    elif raw[:3] == b"\xef\xbb\xbf":
        encoding = "utf-8-sig"

    text = raw.decode(encoding, errors="replace")
    # 去除可能残留的 BOM 字符
    if text and text[0] == "\ufeff":
        text = text[1:]
    # 统一换行符
    text = text.replace("\r\n", "\n").replace("\r", "\n")

    # 自动检测分隔符：tab vs comma
    first_line = text.split("\n")[0] if text else ""
    if first_line.count("\t") >= first_line.count(","):
        delimiter = "\t"
    else:
        delimiter = ","

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = [r for r in reader]
    result: Dict[str, Tuple[tuple, list]] = {}
    sheet_name = os.path.splitext(os.path.basename(filepath))[0]
    parsed = _parse_sheet_rows(rows)
    if parsed is not None:
        result[sheet_name] = parsed
    return result


def read_all_sheets(filepath: str) -> Dict[str, Tuple[tuple, list]]:
    result: Dict[str, Tuple[tuple, list]] = {}
    ext = os.path.splitext(filepath)[1].lower()
    ftype = _detect_file_type(filepath)

    if ftype == "xlsx":
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]
            rows = list(ws.iter_rows(values_only=True))
            parsed = _parse_sheet_rows(rows)
            if parsed is not None:
                result[sname] = parsed
        wb.close()
    elif ftype == "xls":
        wb = xlrd.open_workbook(filepath)
        datemode = wb.datemode
        for sname in wb.sheet_names():
            sheet = wb.sheet_by_name(sname)
            rows = []
            for r in range(sheet.nrows):
                row = []
                for c in range(sheet.ncols):
                    ctype = sheet.cell_type(r, c)
                    val = sheet.cell_value(r, c)
                    if ctype == xlrd.XL_CELL_DATE:
                        try:
                            val = xlrd.xldate_as_datetime(val, datemode)
                        except Exception:
                            pass
                    row.append(val)
                rows.append(row)
            parsed = _parse_sheet_rows(rows)
            if parsed is not None:
                result[sname] = parsed
    elif ftype in ("tsv", "csv"):
        result = _read_text_table(filepath)
    else:
        # 最后兜底：尝试 openpyxl，失败则尝试文本方式
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            for sname in wb.sheetnames:
                ws = wb[sname]
                rows = list(ws.iter_rows(values_only=True))
                parsed = _parse_sheet_rows(rows)
                if parsed is not None:
                    result[sname] = parsed
            wb.close()
        except Exception:
            result = _read_text_table(filepath)
    return result



# 结果产物 sheet 名称集合 — 这些 sheet 是合并输出，不参与二次合并
RESULT_SHEET_NAMES = {
    "全量数据", "交货汇总", "交货汇总_文本日期",
    "数据透析", "工厂交货透视", "奥妙明细", "奥妙小计",
}


def select_source_sheets(sheets_map: Dict[str, Dict]) -> list:
    """从 sheets_map 中自动选择应参与合并的数据源 sheet。

    选择规则：
    1. 排除结果产物 sheet（RESULT_SHEET_NAMES）
    2. 选中明细型 sheet（表头含"交货"列且行数 > 100）
    3. 选中已发运/未发运 sheet（sheet名含"发运"）
    4. 如果以上都没命中，回退选非结果产物且行数 > 50 的 sheet
    5. 如果仍无选中，抛出 ValueError

    Args:
        sheets_map: {sheet_key: {filename, sheet_name, headers, row_count}}

    Returns:
        选中的 sheet key 列表
    """
    selected_keys = []
    for key, info in sheets_map.items():
        sn = info["sheet_name"]
        # 排除结果产物 sheet
        if sn in RESULT_SHEET_NAMES:
            continue
        # 选中明细型 sheet（含"交货"列且行数多）
        headers_str = [str(h) if h else "" for h in info["headers"]]
        has_jiaohuo = any("交货" in h for h in headers_str)
        if has_jiaohuo and info["row_count"] > 100:
            selected_keys.append(key)
        # 选中已发运/未发运
        elif "发运" in sn:
            selected_keys.append(key)

    if not selected_keys:
        # 回退：选非结果产物且行数 > 50 的 sheet
        for key, info in sheets_map.items():
            if info["sheet_name"] in RESULT_SHEET_NAMES:
                continue
            if info["row_count"] > 50:
                selected_keys.append(key)

    if not selected_keys:
        raise ValueError(
            "未找到可合并的数据 sheet（需要总表含「明细」「已发运」「未发运」）"
        )

    return selected_keys


def merge_files(
    file_paths: List[str],
    selected_sheets: Optional[List[str]] = None,
    provinces: Optional[List[str]] = None,
    rule_id: Optional[str] = None,
    output_dir: str = "output",
    output_prefix: str = "合并结果",
    manual_mappings: Optional[Dict] = None,
    date_str: Optional[str] = None,
    delivery_range: Optional[Tuple[int, int]] = None,
) -> Dict:
    """合并多个 Excel 文件为统一标准列，可选按省份/交货号区间筛选，输出 Excel。

    selected_sheets: sheet key 列表，格式 f"{文件名}::{sheet名}"；None 表示全选。
    provinces: 省份列表；None/[] 表示不筛选（全量）。
    rule_id: 规则 id；None 使用内置默认规则。
    manual_mappings: 手动列名映射 {sheet_key: {原始列名: 标准列名}}；None 表示仅自动匹配。
    delivery_range: 可选，交货号区间 (min, max)，如 (2424796922, 2424802864)。
    返回 {output_path, stats, previews}。
    """
    os.makedirs(output_dir, exist_ok=True)
    prov_list = provinces or []

    # 排除不参与合并的文件：下单计划（数据重复，工厂值未映射）
    EXCLUDE_KEYWORDS = ["下单计划"]
    filtered_paths = [
        fp for fp in file_paths
        if not any(kw in os.path.basename(fp) for kw in EXCLUDE_KEYWORDS)
    ]

    # 按文件优先级排序：06o/分销报表/分销下单量/跑单明细 优先
    def _file_priority(fp):
        fn = os.path.basename(fp).lower()
        if "06o" in fn:
            return 0
        if "分销报表" in fn:
            return 1
        if "分销下单量" in fn or "分销-下单量" in fn:
            return 2
        if "跑单明细" in fn:
            return 3
        return 5

    sorted_paths = sorted(filtered_paths, key=_file_priority)
    files_data = {}
    for fp in sorted_paths:
        fname = os.path.basename(fp)
        files_data[fname] = read_all_sheets(fp)

    active_rule = BUILTIN_RULE
    std_header_order = [sh["name"] for sh in BUILTIN_RULE["standard_headers"] if sh.get("name", "").strip()]
    if rule_id:
        for r in load_rules():
            if r["id"] == rule_id:
                active_rule = r
                std_header_order = [sh["name"] for sh in r["standard_headers"] if sh.get("name", "").strip()]
                break

    # selected_sheets 为 None 时，按表头与规则标准表头的匹配率自动筛选 sheet
    # 只选匹配率 >= 60% 的 sheet，排除无关 sheet（如 OMR 导出、预报量等）
    SHEET_MATCH_THRESHOLD = 0.75
    if selected_sheets is None:
        selected_set = set()
        std_header_count = len(std_header_order)
        for fname, sheets in files_data.items():
            for sname, (headers, data_rows) in sheets.items():
                if not headers or not data_rows:
                    continue
                auto_map = match_columns_to_rule(list(headers), active_rule)
                matched_count = sum(1 for m in auto_map if m and m[1])
                match_ratio = matched_count / std_header_count if std_header_count else 0
                if match_ratio >= SHEET_MATCH_THRESHOLD:
                    selected_set.add(f"{fname}::{sname}")
    else:
        selected_set = set(selected_sheets)

    std_col_set = set(std_header_order)
    all_columns = list(std_header_order)

    # 合并所有数据行
    merged_rows = []
    for fname, sheets in files_data.items():
        for sname, (headers, data_rows) in sheets.items():
            key = f"{fname}::{sname}"
            if key not in selected_set:
                continue
            auto_map_list = match_columns_to_rule(headers, active_rule)
            manual_map = (manual_mappings or {}).get(key, {})
            mapped_headers = []
            assigned_std = set()
            for i, h in enumerate(headers):
                hs = str(h) if h else ""
                manual_target = manual_map.get(hs, "")
                std_name = None
                if manual_target and manual_target in std_col_set and manual_target not in assigned_std:
                    std_name = manual_target
                if not std_name and i < len(auto_map_list):
                    auto_std = auto_map_list[i][1] if auto_map_list[i] else None
                    if auto_std and auto_std in std_col_set and auto_std not in assigned_std:
                        std_name = auto_std
                if not std_name and hs in std_col_set and hs not in assigned_std:
                    std_name = hs
                if std_name:
                    mapped_headers.append(std_name)
                    assigned_std.add(std_name)
                else:
                    mapped_headers.append(None)
            for row in data_rows:
                row_dict = {}
                for idx, h in enumerate(mapped_headers):
                    if h is None:
                        continue
                    row_dict[h] = row[idx] if idx < len(row) else None
                for sh in active_rule.get("standard_headers", []):
                    vm = sh.get("value_mappings")
                    if vm:
                        apply_value_mappings(row_dict, sh["name"], vm, fname)
                # 分销报表特殊处理：如果有「提货仓库」列，用提货仓库值覆盖工厂值
                # （提货仓库=YG → 工厂=YG，而非 8136→901 的默认映射）
                if "分销报表" in fname:
                    _thck_idx = None
                    for _i, _h in enumerate(headers):
                        if str(_h).strip() == "提货仓库":
                            _thck_idx = _i
                            break
                    if _thck_idx is not None and _thck_idx < len(row):
                        _thck_val = row[_thck_idx]
                        if _thck_val is not None and str(_thck_val).strip():
                            row_dict["工厂"] = str(_thck_val).strip()
                # 单位转换：统一总重量为吨、业务量为M3
                _wun = str(row_dict.get("WUn", "")).strip()
                if _wun == "公斤":
                    _zw = row_dict.get("总重量")
                    if isinstance(_zw, (int, float)) and _zw:
                        row_dict["总重量"] = _zw / 1000
                    row_dict["WUn"] = "吨"
                _vun = str(row_dict.get("VUn", "")).strip()
                if _vun == "CCM":
                    _yw = row_dict.get("业务量")
                    if isinstance(_yw, (int, float)) and _yw:
                        row_dict["业务量"] = _yw / 1000000
                    row_dict["VUn"] = "M3"
                row_dict["_source_file"] = f"{fname}::{sname}"
                merged_rows.append(row_dict)

    # 对齐 + 去重 + 过滤非法交货号
    aligned = []
    seen_keys = {}  # dedup_key → index in aligned (dict, for field completion)
    for row in merged_rows:
        aligned_row = {col: (row.get(col) if row.get(col) is not None else "") for col in all_columns}
        aligned_row["_source_file"] = row.get("_source_file", "")
        jh_val = str(aligned_row.get("交货", "")).strip()
        if jh_val and not jh_val.isdigit():
            continue
        if not jh_val:
            continue
        xm_val = str(aligned_row.get("项目", "")).strip()
        # 标准化项目号：去掉前导零，使 "000010" 和 "10" 被识别为同一项目
        # 避免不同来源文件（分销报表 vs 跨仓订单）的重复行被重复计算
        if xm_val and xm_val.isdigit():
            xm_normalized = str(int(xm_val))
        else:
            xm_normalized = xm_val
        # 项目号为空时用交货号作为去重键（保留项目号缺失的行）
        dedup_key = (jh_val, xm_normalized) if xm_normalized else (jh_val, "")
        if dedup_key in seen_keys:
            # 去重时做字段补全：用当前行的非空值补到已有行的空字段上
            # 解决不同数据源（如邮件产物 vs 总表）字段完整度不同的问题
            existing = aligned[seen_keys[dedup_key]]
            for col in all_columns:
                if col == "_source_file":
                    continue
                existing_val = existing.get(col, "")
                new_val = aligned_row.get(col, "")
                if (not existing_val or existing_val == "") and new_val:
                    existing[col] = new_val
            continue
        seen_keys[dedup_key] = len(aligned)
        aligned.append(aligned_row)

    # 省份筛选
    if prov_list:
        filtered = [row for row in aligned if match_row_province(row, prov_list)]
    else:
        filtered = aligned

    # 交货号区间筛选（可选，如只取 08-17 批次）
    if delivery_range:
        jhd_min, jhd_max = delivery_range
        def _in_range(row):
            jhd = row.get("交货", "")
            if jhd is None or str(jhd).strip() == "":
                return False
            try:
                v = int(str(jhd).strip())
                return jhd_min <= v <= jhd_max
            except (ValueError, TypeError):
                return False
        filtered = [row for row in filtered if _in_range(row)]

    street_key = None
    for col in all_columns:
        if "街道" in col and "街道2" not in col and "街道 3" not in col:
            street_key = col
            break

    # 构建多 Sheet 输出
    wb = openpyxl.Workbook()
    output_headers = ["售达方" if h == "售达方的名字" else h for h in all_columns]

    ws1 = wb.active
    ws1.title = "全量数据"
    ws1.append(output_headers)
    for row in aligned:
        ws1.append([_excel_date(row.get(h, "")) for h in all_columns])

    ws3 = wb.create_sheet("筛选数据")
    ws3.append(output_headers)
    for row in filtered:
        ws3.append([_excel_date(row.get(h, "")) for h in all_columns])

    # 读取物流信息（已发运/未发运的 B_ADDRESS1/备注，通过销售凭证关联）
    logistics_map = read_logistics_map(files_data)
    # 读取 SO 文件的 NOTES2(→备注)/B_ADDRESS1，通过交货号(=客户订单号)关联
    so_map = read_so_map(files_data)
    # 读取跨仓订单文件的 客户订单号(→B_ADDRESS1)/备注/仓库备注，通过交货号(=OBD)关联
    kuacang_map = read_kuacang_map(files_data)

    p4_headers, p4_data, text_dates = build_pivot_by_delivery(filtered, logistics_map, so_map, kuacang_map)
    ws4 = wb.create_sheet("交货汇总")
    ws4.append(p4_headers)
    for row in p4_data:
        ws4.append(row)

    # 数据透析 sheet（交货汇总的重排+精简版本，按交货号排序）
    dp_headers, dp_data = build_data_pivot(p4_headers, p4_data)
    ws_dp = wb.create_sheet("数据透析")
    ws_dp.append([None] * 12)  # 空行1
    ws_dp.append([None] * 12)  # 空行2
    ws_dp.append(dp_headers)   # 表头
    for row in dp_data:
        ws_dp.append(row)

    p5_headers = list(p4_headers)
    ws5 = wb.create_sheet("交货汇总_文本日期")
    ws5.append(p5_headers)
    for row in p4_data:
        new_row = list(row)
        # 交货 在 index 5
        delivery = str(new_row[5]).strip() if len(new_row) > 5 and new_row[5] else ""
        if delivery == "总计":
            pass  # 总计行保持原样
        else:
            orig = text_dates.get(delivery, {})
            # 发货日期 在 index 0
            if len(new_row) > 0 and new_row[0] is None and orig.get("发货日期"):
                new_row[0] = _format_date_text(orig["发货日期"])
            # 交货日期 在 index 1
            if len(new_row) > 1 and new_row[1] is None and orig.get("交货日期"):
                new_row[1] = _format_date_text(orig["交货日期"])
            # 街道 在 index 7
            if len(new_row) > 7 and new_row[7] is None and orig.get("街道"):
                new_row[7] = orig["街道"]
        ws5.append(new_row)

    p2_headers, p2_data = build_pivot_by_factory_delivery(aligned)
    ws2 = wb.create_sheet("工厂交货透视")
    for row in p2_data:
        ws2.append(row)

    # 生成"奥妙+洗衣粉/皂粉"明细子集和小计表
    omo_detail_rows, omo_subtotal_rows = build_omo_detail(filtered, all_columns)
    if omo_detail_rows:
        ws_omo_detail = wb.create_sheet("奥妙明细")
        ws_omo_detail.append(output_headers)
        for row in omo_detail_rows:
            ws_omo_detail.append([_excel_date(v) for v in row])

        ws_omo_subtotal = wb.create_sheet("奥妙小计")
        ws_omo_subtotal.append(["", "", ""])  # 前2行空
        ws_omo_subtotal.append(["", "", ""])
        ws_omo_subtotal.append(["交货", "求和项:总重量", "求和项:业务量"])
        for row in omo_subtotal_rows:
            ws_omo_subtotal.append(row)

    day = date_str.replace("-", "") if date_str else now_cn().strftime("%Y%m%d")
    short_hash = hashlib.md5(f"{day}_{len(filtered)}_{now_cn().strftime('%H%M%S%f')}".encode()).hexdigest()[:8]
    prov_short = "_".join(p.replace("省", "").replace("市", "") for p in prov_list[:3]) if prov_list else "全部"
    output_filename = f"{output_prefix}_{prov_short}_{day}_{short_hash}.xlsx"
    output_path = os.path.join(output_dir, output_filename)

    # 统一日期格式 + 列宽
    normalize_date_formats(wb)
    wb.save(output_path)
    wb.close()

    stats = {
        "total_merged_rows": len(merged_rows),
        "total_columns": len(all_columns),
        "filtered_rows": len(filtered),
        "pivot_delivery_count": len(p4_data),
        "pivot_factory_count": len(p2_data),
        "omo_detail_count": len(omo_detail_rows),
        "omo_subtotal_count": len(omo_subtotal_rows),
        "street_column": street_key if street_key else "未找到",
        "provinces": prov_list,
    }

    previews = []
    filter_preview_rows = [[row.get(h, "") for h in all_columns] for row in filtered[:PREVIEW_MAX_ROWS]]
    if filter_preview_rows:
        previews.append({
            "sheet_name": "筛选数据",
            "headers": output_headers,
            "rows": [[serialize_cell(c) for c in r] for r in filter_preview_rows],
            "total": len(filtered),
            "preview_count": len(filter_preview_rows),
        })
    previews.append({
        "sheet_name": "交货汇总",
        "headers": [str(h) for h in p4_headers],
        "rows": [[serialize_cell(c) for c in row] for row in p4_data[:PREVIEW_MAX_ROWS]],
        "total": len(p4_data),
        "preview_count": min(len(p4_data), PREVIEW_MAX_ROWS),
    })
    previews.append({
        "sheet_name": "工厂交货透视",
        "headers": [str(h) for h in p2_headers],
        "rows": [[serialize_cell(c) for c in row] for row in p2_data[:PREVIEW_MAX_ROWS]],
        "total": len(p2_data),
        "preview_count": min(len(p2_data), PREVIEW_MAX_ROWS),
    })

    return {"output_path": output_path, "stats": stats, "previews": previews}



# ---- 未发运 sheet 重排序辅助 ----

import re as _re

def _normalize_address_for_sort(address: str) -> str:
    """归一化地址用于排序和分组。

    修正"湖南省长沙县" → "湖南省长沙市长沙县"等格式不一致问题：
    如果"省"后面直接跟"县"（缺少"市"），插入对应地级市名。
    """
    if not address:
        return ""
    addr = address.strip()
    # 匹配 "湖南省XX县" → "湖南省XX市XX县"（XX县 省略了市名）
    m = _re.match(r'^(湖南省)([^市]+县)', addr)
    if m:
        prefix = m.group(1)
        county = m.group(2)
        # 长沙县 → 长沙市长沙县
        city = county.replace('县', '市')
        addr = prefix + city + county + addr[len(m.group(0)):]
    return addr


def _is_subtotal_row(ws, r: int) -> bool:
    """检查第 r 行是否为 SUBTOTAL 公式行（col12 以 =SUBTOTAL 开头）。"""
    c12 = ws.cell(r, 12).value
    return isinstance(c12, str) and c12.startswith("=SUBTOTAL")


def _is_jd_nonbom(b_address1_val) -> bool:
    """判断 B_ADDRESS1 值是否以「京东NONBOM组套订单」开头。"""
    if b_address1_val is None:
        return False
    return str(b_address1_val).startswith("京东NONBOM组套订单")


def _restructure_weifayun_sheet(ws, factory_fallback_fill, std_border, red_font):
    """重排未发运 sheet。

    规则：
    - 901 工厂行永远在最上面（含 B_ADDRESS1 京东NONBOM 强制改为 901 的行）
    - 每个区内按「客户名称」(col10) 分组，每组末尾插 SUBTOTAL 小计行
    - 同一客户分组内，同工厂值的行聚拢不交错（按 factory + 原始行号稳定排序）
    - B_ADDRESS1 以「京东NONBOM组套订单」开头 → 工厂强制改 901，B_ADDRESS1 加粗棕色
    - 空客户名称行排到数据区末尾，不加小计
    """
    from copy import copy as _copy
    from collections import OrderedDict
    from openpyxl.styles import PatternFill, Font, Alignment, Border

    max_col = ws.max_column
    _default_font = Font(size=11)
    _jd_font = Font(color="FF806000", bold=True)  # 棕色加粗

    # ---- 1. 全表扫描收集 + 京东NONBOM识别 ----
    all_rows = []  # [{vals, is_jd, orig_idx}]
    for r in range(2, ws.max_row + 1):
        if _is_subtotal_row(ws, r):
            continue
        # 跳过空行（关键列全空）
        if all(ws.cell(r, c).value is None for c in (3, 4, 5, 6, 10, 12, 15)):
            continue
        row_vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        # 京东NONBOM 识别：强制改工厂为 901
        is_jd = _is_jd_nonbom(row_vals[20] if len(row_vals) > 20 else None)
        if is_jd:
            row_vals[14] = "901"
        all_rows.append({"vals": row_vals, "is_jd": is_jd, "orig_idx": r})

    # ---- 2. 901区/非901区拆分 ----
    rows_901 = []
    rows_non_901 = []
    for row in all_rows:
        factory = str(row["vals"][14]).strip() if len(row["vals"]) > 14 and row["vals"][14] else ""
        if factory == "901":
            rows_901.append(row)
        else:
            rows_non_901.append(row)

    def _group_by_customer(rows):
        """按客户名称(col10, index9)分组，空客户名称行不归入。返回 (分组, 空客户行列表)。"""
        groups = OrderedDict()
        no_customer = []
        for row in rows:
            vals = row["vals"]
            cust = str(vals[9]).strip() if len(vals) > 9 and vals[9] else ""
            if not cust:
                no_customer.append(row)
                continue
            if cust not in groups:
                groups[cust] = []
            groups[cust].append(row)
        return groups, no_customer

    groups_901, no_cust_901 = _group_by_customer(rows_901)
    groups_non, no_cust_non = _group_by_customer(rows_non_901)
    no_customer_rows = no_cust_901 + no_cust_non

    # ---- 3. 清空数据区（保留表头）----
    for r in range(2, ws.max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            cell.value = None
            cell.fill = PatternFill()
            cell.font = Font()
            cell.border = Border()
            cell.alignment = Alignment()
            cell.number_format = "General"

    # ---- 4. 逐组写入 + SUBTOTAL ----
    current_row = 2

    def _write_data_row(row):
        nonlocal current_row
        vals = row["vals"]
        status = str(vals[6]).strip() if len(vals) > 6 and vals[6] else ""
        factory = str(vals[14]).strip() if len(vals) > 14 and vals[14] else ""

        for c in range(1, max_col + 1):
            cell = ws.cell(current_row, c)
            val = vals[c - 1] if c - 1 < len(vals) else None
            cell.value = val
            cell.border = _copy(std_border)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.font = _copy(_default_font)

        # 工厂填充色
        canonical_fill = factory_fallback_fill.get(factory)
        if canonical_fill:
            ws.cell(current_row, 15).fill = _copy(canonical_fill)

        # 不可提 → col7 红色加粗
        if status == "不可提":
            ws.cell(current_row, 7).font = _copy(red_font)

        # 京东NONBOM → col21 棕色加粗
        if row["is_jd"]:
            ws.cell(current_row, 21).font = _copy(_jd_font)

        current_row += 1

    def _write_group(groups):
        nonlocal current_row
        for cust, rows in groups.items():
            # 组内按 (factory, orig_idx) 稳定排序聚拢
            rows.sort(key=lambda row: (
                str(row["vals"][14]).strip() if len(row["vals"]) > 14 and row["vals"][14] else "",
                row["orig_idx"],
            ))
            group_start = current_row
            for row in rows:
                _write_data_row(row)
            # 组末尾 SUBTOTAL
            _write_wfy_subtotal(ws, current_row, group_start, current_row - 1,
                                max_col, std_border, red_font)
            current_row += 1

    # 901区各组 → 非901区各组 → 空客户行（不加小计）
    _write_group(groups_901)
    _write_group(groups_non)
    for row in no_customer_rows:
        _write_data_row(row)

    # ---- 5. 清除尾部多余空行 ----
    if current_row - 1 < ws.max_row:
        ws.delete_rows(current_row, ws.max_row - (current_row - 1))

    return current_row - 2  # 总写入行数（含 SUBTOTAL）


def _write_wfy_subtotal(ws, row, start, end, max_col, std_border, red_font):
    """在指定行写入 SUBTOTAL 公式行（L/M/N 列），并应用样式。"""
    from copy import copy as _copy
    from openpyxl.styles import Alignment

    ws.cell(row, 12, value=f"=SUBTOTAL(9,L{start}:L{end})")
    ws.cell(row, 13, value=f"=SUBTOTAL(9,M{start}:M{end})")
    ws.cell(row, 14, value=f"=SUBTOTAL(9,N{start}:N{end})")
    for col in range(1, max_col + 1):
        cell = ws.cell(row, col)
        cell.border = _copy(std_border)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        if col in (12, 13, 14):
            cell.font = _copy(red_font)



def merge_mail_into_master(master_path: str, mail_path: str, output_path: str = None) -> dict:
    """将邮件捞取的每日数据追加到总表（master table）格式中。

    保留总表的原始格式（已发运/未发运/明细/客户信息/组套/Sheet5）：
    - 明细 sheet：追加筛选数据中所有新交货号的所有行（同一交货号多行全部追加）
    - 未发运 sheet：从交货汇总中追加新订单（B_ADDRESS1/备注交换）
    - 已发运、Sheet5、客户信息、组套：不修改

    Args:
        master_path: 总表文件路径（上传的 xlsx）
        mail_path: 邮件捞取产物文件路径
        output_path: 输出文件路径（可选，默认自动生成）

    Returns:
        dict: {"output_path": str, "appended_count": int, "total_in_detail": int,
               "appended_weifayun_count": int}
    """
    import shutil

    if output_path is None:
        output_path = os.path.join(config.OUTPUT_DIR, f"总表合并_{fromtimestamp_cn(datetime.datetime.now().timestamp()).strftime('%Y%m%d_%H%M%S')}.xlsx")

    # 复制总表作为输出基础（保留所有 sheet 和格式）
    shutil.copy2(master_path, output_path)
    # 确保输出文件可写（源文件可能来自微信附件，是只读的）
    os.chmod(output_path, 0o644)

    # 读取邮件捞取产物（先读取，后面多处使用）
    mail_wb = openpyxl.load_workbook(mail_path, read_only=True, data_only=True)

    # ---- 1. 读取筛选数据 ----
    source_sheet = None
    for sn in ("筛选数据", "全量数据"):
        if sn in mail_wb.sheetnames:
            source_sheet = sn
            break
    if source_sheet is None:
        mail_wb.close()
        raise ValueError("邮件捞取产物中没有「筛选数据」或「全量数据」sheet")

    ws_mail = mail_wb[source_sheet]
    mail_rows = list(ws_mail.iter_rows(min_row=2, values_only=True))

    # ---- 2. 读取交货汇总 ----
    summary_rows = []
    summary_headers = []
    if "交货汇总" in mail_wb.sheetnames:
        ws_summary = mail_wb["交货汇总"]
        summary_headers = [c.value for c in ws_summary[1]]
        summary_rows = list(ws_summary.iter_rows(min_row=2, values_only=True))
    mail_wb.close()

    # 按表头名定位 备注/B_ADDRESS1 列（兼容新旧两种列序的邮件捞取产物；
    # 找不到表头时回退到新列序的固定位置 12/13）
    def _summary_col(name, fallback):
        for _i, _h in enumerate(summary_headers):
            if _h is not None and str(_h).strip() == name:
                return _i
        return fallback

    ci_beizhu = _summary_col("备注", 12)
    ci_baddress = _summary_col("B_ADDRESS1", 13)

    # ---- 3. 打开输出文件，准备追加 ----
    master_wb = openpyxl.load_workbook(output_path)
    if "明细" not in master_wb.sheetnames:
        master_wb.close()
        raise ValueError("总表中没有「明细」sheet")

    # 读取总表明细 sheet 中已有的交货号集合
    ws_detail = master_wb["明细"]
    existing_jiaohuo = set()
    max_row = 1  # header row
    for row in ws_detail.iter_rows(min_row=2, values_only=False):
        val = row[0].value if row else None
        if val is not None and str(val).strip():
            existing_jiaohuo.add(str(val).strip())
            max_row = row[0].row

    # 读取总表已发运/未发运中已有的订单号集合（用于未发运去重）
    existing_weifayun_orders = set()
    if "未发运" in master_wb.sheetnames:
        for row in master_wb["未发运"].iter_rows(min_row=2, values_only=True):
            if row and len(row) > 5 and row[5] is not None:
                existing_weifayun_orders.add(str(row[5]).strip())

    existing_yifayun_orders = set()
    if "已发运" in master_wb.sheetnames:
        for row in master_wb["已发运"].iter_rows(min_row=2, values_only=True):
            if row and len(row) > 5 and row[5] is not None:
                existing_yifayun_orders.add(str(row[5]).strip())

    # ---- 4. 追加明细 sheet（所有新交货号的所有行）----
    # 总表明细有 37 列，邮件数据有 34 列
    # 列映射：总表[0]=交货, 总表[1]=空, 总表[2..34]=邮件[1..33], 总表[35..36]=空
    # 样式：从总表已有行中找到相同工厂值的行作为模板，复制其 fill/font/alignment/number_format
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.styles.colors import Color
    # 标准细边框（田字格），用于无模板行时
    _thin = Side(style="thin")
    _STD_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    # 建立 工厂值 → 模板行号 映射（工厂在第13列/M）
    # 未发运库区(工厂) → 填充色 的标准回退映射（当总表无对应工厂模板行时使用）
    # 基于参考总表的实际颜色
    _WFY_FACTORY_FALLBACK_FILL = {
        "901":  PatternFill(patternType="solid", fgColor="FFFF0000"),   # 红色
        "801":  PatternFill(patternType="solid", fgColor="FF92D050"),   # 绿色
        "8137": PatternFill(patternType="solid", fgColor="FFFFFF00"),   # 黄色
        "8205": PatternFill(patternType="solid", fgColor=Color(theme=4, tint=0.6)),  # 主题色
        "301":  PatternFill(patternType="solid", fgColor="FF92D050"),   # 绿色
        "YG":   PatternFill(patternType="solid", fgColor="FFFFFF00"),   # 黄色
        "701":  PatternFill(patternType="solid", fgColor="FFFFFF00"),   # 黄色
        # 8136: 无填充（保持默认）
    }
    detail_factory_template = {}  # factory_str -> row_num
    for r in range(2, max_row + 1):
        fv = ws_detail.cell(row=r, column=13).value
        if fv is not None and str(fv) not in detail_factory_template:
            detail_factory_template[str(fv)] = r
    appended_count = 0
    for mail_row in mail_rows:
        if not mail_row or mail_row[0] is None:
            continue
        jiaohuo = str(mail_row[0]).strip()
        if not jiaohuo or jiaohuo in existing_jiaohuo:
            continue

        # 新交货号，追加到明细 sheet（不加入 existing_jiaohuo，允许同交货号多行追加）
        max_row += 1
        ws_detail.cell(row=max_row, column=1, value=mail_row[0])
        # col B: VLOOKUP 公式（与原有一致）
        ws_detail.cell(row=max_row, column=2, value=f"=VLOOKUP(A{max_row},Sheet5!A:C,3,0)")
        for mail_col in range(1, min(34, len(mail_row))):
            master_col = mail_col + 2
            val = mail_row[mail_col]
            if val is not None:
                ws_detail.cell(row=max_row, column=master_col, value=val)

        # 复制样式：找到同工厂值的模板行，逐列复制 fill/font/alignment/number_format
        factory_val = ws_detail.cell(row=max_row, column=13).value
        template_row = detail_factory_template.get(str(factory_val)) if factory_val else None
        if template_row:
            for col in range(1, ws_detail.max_column + 1):
                t_cell = ws_detail.cell(row=template_row, column=col)
                n_cell = ws_detail.cell(row=max_row, column=col)
                if t_cell.has_style:
                    n_cell.fill = copy(t_cell.fill)
                    n_cell.font = copy(t_cell.font)
                    n_cell.alignment = copy(t_cell.alignment)
                    n_cell.number_format = t_cell.number_format
                    n_cell.border = copy(t_cell.border)

        appended_count += 1

    # ---- 5. 追加未发运 sheet（从交货汇总，B_ADDRESS1/备注交换）----
    # 新行按到货城市分组，每个城市连续放在一起，城市间用空行分隔
    appended_weifayun_count = 0
    if summary_rows and "未发运" in master_wb.sheetnames:
        ws_weifayun = master_wb["未发运"]
        # 找到未发运最后一行
        wfy_max_row = 1
        for row in ws_weifayun.iter_rows(min_row=2, values_only=False):
            if any(c.value is not None for c in row):
                wfy_max_row = row[0].row
        # 建立 库区(工厂)值 → 模板行号 映射（库区在第15列/O）
        wfy_factory_template = {}  # factory_str -> row_num
        for r in range(2, wfy_max_row + 1):
            fv = ws_weifayun.cell(row=r, column=15).value
            if fv is not None and str(fv) not in wfy_factory_template:
                wfy_factory_template[str(fv)] = r

        # 先收集所有新行数据（尚未写入），按城市分组
        new_rows_by_city = {}  # city_str -> list of s_row
        city_order = []  # 保持城市首次出现顺序
        for s_row in summary_rows:
            if not s_row or len(s_row) < 6:
                continue
            # 跳过总计行：col0 为"总计"或 col5（交货）为空
            if s_row[5] is None or not str(s_row[5]).strip():
                continue
            if s_row[0] is not None and str(s_row[0]).strip() == "总计":
                continue

            jiaohuo = str(s_row[5]).strip()
            # 跳过已在未发运或已发运中的订单号
            if jiaohuo in existing_weifayun_orders or jiaohuo in existing_yifayun_orders:
                continue

            city = str(s_row[2]).strip() if s_row[2] is not None else ""
            if city not in new_rows_by_city:
                new_rows_by_city[city] = []
                city_order.append(city)
            new_rows_by_city[city].append(s_row)
            existing_weifayun_orders.add(jiaohuo)

        # 按城市分组写入，每个城市组末尾添加 SUBTOTAL 小计行（红色字体）
        _RED_FONT = Font(color="FFFF0000", bold=True)
        for city in city_order:
            group_start_row = wfy_max_row + 1
            for s_row in new_rows_by_city[city]:
                wfy_max_row += 1
                # 字段映射（交货汇总 → 未发运），B_ADDRESS1/备注交换
                # col0 发货日期 → col0 下单日期
                ws_weifayun.cell(row=wfy_max_row, column=1, value=s_row[0])
                # col1 交货日期 → col1 需求日期
                ws_weifayun.cell(row=wfy_max_row, column=2, value=s_row[1])
                # col2 送达方地点 → col2 到货城市
                ws_weifayun.cell(row=wfy_max_row, column=3, value=s_row[2])
                # col3 运达方 → col3 客户代码
                ws_weifayun.cell(row=wfy_max_row, column=4, value=s_row[3])
                # col4 销售凭证 → col4 销售凭证
                ws_weifayun.cell(row=wfy_max_row, column=5, value=s_row[4])
                # col5 交货 → col5 订单号
                ws_weifayun.cell(row=wfy_max_row, column=6, value=s_row[5])
                # col9 客户名称 = 交货汇总 col6 运达方的名字
                ws_weifayun.cell(row=wfy_max_row, column=10, value=s_row[6])
                # col10 客户地址 = 交货汇总 col7 街道
                ws_weifayun.cell(row=wfy_max_row, column=11, value=s_row[7])
                # col11 数量 = 交货汇总 col8 求和项:交货量
                ws_weifayun.cell(row=wfy_max_row, column=12, value=s_row[8])
                # col12 吨位 = 交货汇总 col9 求和项:总重量
                ws_weifayun.cell(row=wfy_max_row, column=13, value=s_row[9])
                # col13 体积 = 交货汇总 col10 求和项:业务量
                ws_weifayun.cell(row=wfy_max_row, column=14, value=s_row[10])
                # col14 库区 = 交货汇总 col11 工厂
                ws_weifayun.cell(row=wfy_max_row, column=15, value=s_row[11])
                # 同名直连：总表 B_ADDRESS1 ← 交货汇总「B_ADDRESS1」，总表 备注 ← 交货汇总「备注」
                # （列位置按表头名解析，兼容不同列序的邮件捞取产物）
                ws_weifayun.cell(row=wfy_max_row, column=21, value=s_row[ci_baddress])
                ws_weifayun.cell(row=wfy_max_row, column=22, value=s_row[ci_beizhu])

                # 复制样式：找到同库区(工厂)值的模板行，逐列复制 fill/font/alignment/number_format
                factory_val = ws_weifayun.cell(row=wfy_max_row, column=15).value
                factory_str = str(factory_val).strip() if factory_val else ""
                template_row = wfy_factory_template.get(factory_str) if factory_str else None
                if template_row:
                    for col in range(1, ws_weifayun.max_column + 1):
                        t_cell = ws_weifayun.cell(row=template_row, column=col)
                        n_cell = ws_weifayun.cell(row=wfy_max_row, column=col)
                        if t_cell.has_style:
                            n_cell.fill = copy(t_cell.fill)
                            n_cell.font = copy(t_cell.font)
                            n_cell.alignment = copy(t_cell.alignment)
                            n_cell.number_format = t_cell.number_format
                            n_cell.border = copy(t_cell.border)
                # 无模板行时，应用标准细边框（田字格）+ 左对齐居中
                if not template_row:
                    for col in range(1, ws_weifayun.max_column + 1):
                        ws_weifayun.cell(row=wfy_max_row, column=col).border = copy(_STD_BORDER)
                        ws_weifayun.cell(row=wfy_max_row, column=col).alignment = Alignment(horizontal="left", vertical="center")
                # 统一覆盖 col15(库区) 填充色为标准颜色（以参考总表为准）
                canonical_fill = _WFY_FACTORY_FALLBACK_FILL.get(factory_str)
                if canonical_fill:
                    ws_weifayun.cell(row=wfy_max_row, column=15).fill = copy(canonical_fill)

                appended_weifayun_count += 1
            # SUBTOTAL 由 _restructure_weifayun_sheet 统一生成，此处不重复写

    # ---- 6. 重排未发运 sheet：901保持原始顺序分组，非901按地址排序 + SUBTOTAL ----
    if "未发运" in master_wb.sheetnames:
        _restructure_weifayun_sheet(
            master_wb["未发运"],
            _WFY_FACTORY_FALLBACK_FILL,
            _STD_BORDER,
            Font(color="FFFF0000", bold=True),
        )

    # 统一日期格式 + 列宽
    normalize_date_formats(master_wb)
    master_wb.save(output_path)
    master_wb.close()

    return {
        "output_path": output_path,
        "appended_count": appended_count,
        "total_in_detail": max_row - 1,
        "appended_weifayun_count": appended_weifayun_count,
    }
