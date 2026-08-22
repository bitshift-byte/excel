"""
TDD tests for merge_mail_into_master — 
merges daily mail-fetched data INTO the 总表 (master table) format.

Instead of generating analysis sheets (全量数据/筛选数据/交货汇总),
this function appends new 交货号 from the mail-fetched file
to the 总表's 明细 sheet, preserving the 总表 format.
"""
import os
import datetime
import openpyxl
import pytest
from merger import merge_mail_into_master


# 总表 明细 headers (37 cols, with empty cols at 1, 35, 36)
MASTER_HEADERS = (
    "交货", "", "DlvTy", "项目", "物料", "描述", "存储位置", "销售凭证",
    "运达方", "运达方的名字", "送达方地点", "名称 3", "工厂", "路线", "OPS", "WhN",
    "批次", "仓位", "GM", "销售组织", "售达方", "售达方", "街道", "街道2",
    "街道 3", "交货量", "SU", "数量(库存单位)", "计", "总重量", "WUn",
    "业务量", "VUn", "交货日期", "发货日期", "", "",
)

# Mail-fetched 筛选数据 headers (34 cols)
MAIL_HEADERS = (
    "交货", "DlvTy", "项目", "物料", "描述", "存储位置", "销售凭证",
    "运达方", "运达方的名字", "送达方地点", "名称 3", "工厂", "路线", "OPS", "WhN",
    "批次", "仓位", "GM", "销售组织", "售达方", "售达方", "街道", "街道2",
    "街道 3", "交货量", "SU", "数量(库存单位)", "计", "总重量", "WUn",
    "业务量", "VUn", "交货日期", "发货日期",
)


def _make_master_row(jiaohuo="2424827207"):
    """Build a 明细 row in 总表 format (37 cols)."""
    row = [None] * 37
    row[0] = jiaohuo
    row[2] = "ZLF1"
    row[3] = "000010"
    row[33] = datetime.datetime(2026, 8, 26)
    row[34] = datetime.datetime(2026, 8, 22)
    return tuple(row)


def _make_mail_row(jiaohuo="2424827180"):
    """Build a 筛选数据 row in mail format (34 cols)."""
    row = [None] * 34
    row[0] = jiaohuo
    row[1] = "ZLF1"
    row[2] = "000010"
    row[32] = datetime.datetime(2026, 8, 28)
    row[33] = datetime.datetime(2026, 8, 22)
    return tuple(row)


def _create_master_file(path, existing_rows):
    """Create a 总表 xlsx with 明细 sheet containing existing rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "明细"
    # Write headers
    for c, h in enumerate(MASTER_HEADERS, 1):
        ws.cell(row=1, column=c, value=h)
    # Write existing data
    for r, row in enumerate(existing_rows, 2):
        for c, v in enumerate(row, 1):
            if v is not None:
                ws.cell(row=r, column=c, value=v)
    # Add other sheets
    for sn in ("已发运", "未发运", "客户信息", "组套", "Sheet5"):
        wb.create_sheet(sn)
        ws2 = wb[sn]
        ws2.cell(row=1, column=1, value="header")
    wb.save(path)
    wb.close()


def _create_mail_file(path, rows, sheet_name="筛选数据"):
    """Create a mail-fetched xlsx with 筛选数据 sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    # Write headers
    for c, h in enumerate(MAIL_HEADERS, 1):
        ws.cell(row=1, column=c, value=h)
    # Write data
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            if v is not None:
                ws.cell(row=r, column=c, value=v)
    # Also add 全量数据 sheet (same structure)
    ws2 = wb.create_sheet("全量数据")
    for c, h in enumerate(MAIL_HEADERS, 1):
        ws2.cell(row=1, column=c, value=h)
    wb.save(path)
    wb.close()


# =====================================================================
# RED 1: New 交货号 from mail file appended to 明细
# =====================================================================

def test_merge_appends_new_jiaohuo(tmp_path):
    """New 交货号 from mail file should be appended to 明细 sheet."""
    master_path = str(tmp_path / "master.xlsx")
    mail_path = str(tmp_path / "mail.xlsx")
    output_path = str(tmp_path / "output.xlsx")

    _create_master_file(master_path, [_make_master_row("2424827207")])
    _create_mail_file(mail_path, [_make_mail_row("2424827180")])

    result = merge_mail_into_master(master_path, mail_path, output_path)

    wb = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
    ws = wb["明细"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    data_rows = [r for r in rows if r[0] is not None]
    
    # Should have 2 rows: original + new
    assert len(data_rows) == 2
    jiaohuo_list = [str(r[0]) for r in data_rows]
    assert "2424827207" in jiaohuo_list
    assert "2424827180" in jiaohuo_list
    wb.close()


# =====================================================================
# RED 2: Duplicate 交货号 NOT re-appended
# =====================================================================

def test_merge_skips_duplicate_jiaohuo(tmp_path):
    """交货号 already in 明细 should not be appended again."""
    master_path = str(tmp_path / "master.xlsx")
    mail_path = str(tmp_path / "mail.xlsx")
    output_path = str(tmp_path / "output.xlsx")

    _create_master_file(master_path, [_make_master_row("2424827207")])
    _create_mail_file(mail_path, [_make_mail_row("2424827207")])  # same 交货号

    merge_mail_into_master(master_path, mail_path, output_path)

    wb = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
    ws = wb["明细"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    data_rows = [r for r in rows if r[0] is not None]
    
    # Should still have only 1 row (no duplicate)
    assert len(data_rows) == 1
    wb.close()


# =====================================================================
# RED 3: All 总表 sheets preserved
# =====================================================================

def test_merge_preserves_all_sheets(tmp_path):
    """Output should have all sheets from 总表: 已发运, 未发运, 明细, etc."""
    master_path = str(tmp_path / "master.xlsx")
    mail_path = str(tmp_path / "mail.xlsx")
    output_path = str(tmp_path / "output.xlsx")

    _create_master_file(master_path, [_make_master_row("2424827207")])
    _create_mail_file(mail_path, [_make_mail_row("2424827180")])

    merge_mail_into_master(master_path, mail_path, output_path)

    wb = openpyxl.load_workbook(output_path, read_only=True)
    expected_sheets = {"明细", "已发运", "未发运", "客户信息", "组套", "Sheet5"}
    assert set(wb.sheetnames) == expected_sheets
    wb.close()


# =====================================================================
# RED 4: Column mapping correct (37 cols, empty at 1, 35, 36)
# =====================================================================

def test_merge_column_mapping(tmp_path):
    """Appended rows should have correct column mapping (37 cols)."""
    master_path = str(tmp_path / "master.xlsx")
    mail_path = str(tmp_path / "mail.xlsx")
    output_path = str(tmp_path / "output.xlsx")

    _create_master_file(master_path, [])
    _create_mail_file(mail_path, [_make_mail_row("2424827180")])

    merge_mail_into_master(master_path, mail_path, output_path)

    wb = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
    ws = wb["明细"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    data_rows = [r for r in rows if r[0] is not None]
    
    assert len(data_rows) == 1
    row = data_rows[0]
    # 37 columns
    assert len(row) == 37
    # 交货号 in col 0
    assert str(row[0]) == "2424827180"
    # Col 1 is empty
    assert row[1] is None or row[1] == ""
    # DlvTy in col 2 (was col 1 in mail)
    assert str(row[2]) == "ZLF1"
    # 项目 in col 3 (was col 2 in mail)
    assert str(row[3]) == "000010"
    # 交货日期 in col 33 (was col 32 in mail)
    assert row[33] is not None
    # 发货日期 in col 34 (was col 33 in mail)
    assert row[34] is not None
    wb.close()


# =====================================================================
# RED 5: Returns output path
# =====================================================================

def test_merge_returns_output_path(tmp_path):
    """Function should return the output file path."""
    master_path = str(tmp_path / "master.xlsx")
    mail_path = str(tmp_path / "mail.xlsx")
    output_path = str(tmp_path / "output.xlsx")

    _create_master_file(master_path, [])
    _create_mail_file(mail_path, [_make_mail_row("2424827180")])

    result = merge_mail_into_master(master_path, mail_path, output_path)
    
    assert "output_path" in result
    assert os.path.exists(result["output_path"])
    assert "appended_count" in result
    assert result["appended_count"] == 1


# =====================================================================
# RED 6: Multiple new 交货号 appended
# =====================================================================

def test_merge_multiple_new(tmp_path):
    """Multiple new 交货号 should all be appended."""
    master_path = str(tmp_path / "master.xlsx")
    mail_path = str(tmp_path / "mail.xlsx")
    output_path = str(tmp_path / "output.xlsx")

    _create_master_file(master_path, [
        _make_master_row("2424827207"),
        _make_master_row("2424827208"),
    ])
    _create_mail_file(mail_path, [
        _make_mail_row("2424827180"),
        _make_mail_row("2424827181"),
        _make_mail_row("2424827182"),
        _make_mail_row("2424827207"),  # duplicate, should be skipped
    ])

    result = merge_mail_into_master(master_path, mail_path, output_path)

    wb = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
    ws = wb["明细"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    data_rows = [r for r in rows if r[0] is not None]
    
    # 2 original + 3 new (1 duplicate skipped)
    assert len(data_rows) == 5
    assert result["appended_count"] == 3
    wb.close()


# =====================================================================
# 未发运 (unshipped) sheet headers — 29 columns
# =====================================================================

WEIFAYUN_HEADERS = (
    "下单日期", "需求日期", "到货城市", "客户代码", "销售凭证", "订单号",
    "提货状态", "拼车", "装车顺序", "客户名称", "客户地址", "数量",
    "吨位", "体积", "库区", "平台/直送", "粉吨位", "粉体积", "司机",
    "车型", "B_ADDRESS1", "备注", "备注-库房属性+直通直送",
    "回单情况", "赠品采购单号", "件数", "箱数", "体积", "重量",
)

# 交货汇总 (delivery summary) headers — 16 columns
JIAOHUO_SUMMARY_HEADERS = (
    "发货日期", "交货日期", "送达方地点", "运达方", "销售凭证",
    "交货", "运达方的名字", "街道",
    "求和项:交货量", "求和项:总重量", "求和项:业务量", "工厂",
    "B_ADDRESS1", "备注", "1", "2",
)


def _make_summary_row(
    jiaohuo="2424827180",
    bao_date=datetime.datetime(2026, 8, 22),
    huo_date=datetime.datetime(2026, 8, 28),
    city="汝城县",
    yundafang="18011338",
    xiaoshou="5507040770",
    name="朝批方盛",
    street="湖南省郴州市",
    qty=32,
    weight=0.16,
    volume=0.521,
    factory="801",
    b_address1="QJDCP20260821BW跨仓03-8271",
    beizhu="26082100001636",
):
    """Build a 交货汇总 row (16 cols)."""
    return (
        bao_date, huo_date, city, yundafang, xiaoshou, jiaohuo,
        name, street, qty, weight, volume, factory,
        b_address1, beizhu, None, None,
    )


def _create_master_with_weifayun(path, detail_rows, weifayun_rows, yifayun_rows=None):
    """Create a 总表 with 明细, 未发运, 已发运 sheets all populated."""
    wb = openpyxl.Workbook()

    # 明细 sheet
    ws = wb.active
    ws.title = "明细"
    for c, h in enumerate(MASTER_HEADERS, 1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(detail_rows, 2):
        for c, v in enumerate(row, 1):
            if v is not None:
                ws.cell(row=r, column=c, value=v)

    # 未发运 sheet
    ws_w = wb.create_sheet("未发运")
    for c, h in enumerate(WEIFAYUN_HEADERS, 1):
        ws_w.cell(row=1, column=c, value=h)
    for r, row in enumerate(weifayun_rows, 2):
        for c, v in enumerate(row, 1):
            if v is not None:
                ws_w.cell(row=r, column=c, value=v)

    # 已发运 sheet
    ws_y = wb.create_sheet("已发运")
    for c, h in enumerate(WEIFAYUN_HEADERS, 1):
        ws_y.cell(row=1, column=c, value=h)
    if yifayun_rows:
        for r, row in enumerate(yifayun_rows, 2):
            for c, v in enumerate(row, 1):
                if v is not None:
                    ws_y.cell(row=r, column=c, value=v)

    # Other sheets
    for sn in ("客户信息", "组套", "Sheet5"):
        wb.create_sheet(sn)
        wb[sn].cell(row=1, column=1, value="header")

    wb.save(path)
    wb.close()


def _create_mail_with_summary(path, filter_rows, summary_rows):
    """Create a mail-fetched xlsx with 筛选数据 + 交货汇总 sheets."""
    wb = openpyxl.Workbook()

    # 筛选数据 sheet
    ws = wb.active
    ws.title = "筛选数据"
    for c, h in enumerate(MAIL_HEADERS, 1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(filter_rows, 2):
        for c, v in enumerate(row, 1):
            if v is not None:
                ws.cell(row=r, column=c, value=v)

    # 全量数据 sheet (same structure, can be empty)
    ws2 = wb.create_sheet("全量数据")
    for c, h in enumerate(MAIL_HEADERS, 1):
        ws2.cell(row=1, column=c, value=h)

    # 交货汇总 sheet
    ws3 = wb.create_sheet("交货汇总")
    for c, h in enumerate(JIAOHUO_SUMMARY_HEADERS, 1):
        ws3.cell(row=1, column=c, value=h)
    for r, row in enumerate(summary_rows, 2):
        for c, v in enumerate(row, 1):
            if v is not None:
                ws3.cell(row=r, column=c, value=v)

    wb.save(path)
    wb.close()


# =====================================================================
# RED 7: 明细 — all rows for same 交货号 appended (not just first)
# =====================================================================

def test_merge_detail_appends_all_rows_same_jiaohuo(tmp_path):
    """Multiple rows with same 交货号 (different 项目/物料) should ALL be appended."""
    master_path = str(tmp_path / "master.xlsx")
    mail_path = str(tmp_path / "mail.xlsx")
    output_path = str(tmp_path / "output.xlsx")

    _create_master_file(master_path, [_make_master_row("2424827207")])

    # Same 交货号, different 项目
    mail_rows = [
        _make_mail_row("2424827180"),  # 项目=000010
        _make_mail_row("2424827180"),  # duplicate 交货号, should still be appended
        _make_mail_row("2424827180"),  # third row, same 交货号
    ]
    _create_mail_file(mail_path, mail_rows)

    result = merge_mail_into_master(master_path, mail_path, output_path)

    wb = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
    ws = wb["明细"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    data_rows = [r for r in rows if r[0] is not None]

    # 1 original + 3 new (all rows, not just first)
    assert len(data_rows) == 4
    assert result["appended_count"] == 3
    wb.close()


# =====================================================================
# RED 8: 未发运 — new orders appended from 交货汇总
# =====================================================================

def test_merge_weifayun_appends_new_orders(tmp_path):
    """交货汇总 with new 交货号 should add rows to 未发运 sheet."""
    master_path = str(tmp_path / "master.xlsx")
    mail_path = str(tmp_path / "mail.xlsx")
    output_path = str(tmp_path / "output.xlsx")

    _create_master_with_weifayun(
        master_path,
        detail_rows=[_make_master_row("2424827207")],
        weifayun_rows=[],
    )

    summary_rows = [
        _make_summary_row(jiaohuo="2424827180"),
        _make_summary_row(jiaohuo="2424827181"),
    ]
    _create_mail_with_summary(mail_path, filter_rows=[], summary_rows=summary_rows)

    result = merge_mail_into_master(master_path, mail_path, output_path)

    wb = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
    ws = wb["未发运"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    data_rows = [r for r in rows if r[5] is not None]

    assert len(data_rows) == 2
    assert "appended_weifayun_count" in result
    assert result["appended_weifayun_count"] == 2
    wb.close()


# =====================================================================
# RED 9: 未发运 — B_ADDRESS1/备注 swapped from 交货汇总
# =====================================================================

def test_merge_weifayun_swaps_b_address1_and_beizhu(tmp_path):
    """交货汇总 B_ADDRESS1→未发运备注, 交货汇总备注→未发运B_ADDRESS1 (swapped)."""
    master_path = str(tmp_path / "master.xlsx")
    mail_path = str(tmp_path / "mail.xlsx")
    output_path = str(tmp_path / "output.xlsx")

    _create_master_with_weifayun(
        master_path,
        detail_rows=[_make_master_row("2424827207")],
        weifayun_rows=[],
    )

    summary_rows = [
        _make_summary_row(
            jiaohuo="2424827180",
            b_address1="SUMMARY_BA_VALUE",
            beizhu="SUMMARY_BZ_VALUE",
        ),
    ]
    _create_mail_with_summary(mail_path, filter_rows=[], summary_rows=summary_rows)

    merge_mail_into_master(master_path, mail_path, output_path)

    wb = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
    ws = wb["未发运"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    data_rows = [r for r in rows if r[5] is not None]

    assert len(data_rows) == 1
    row = data_rows[0]
    # B_ADDRESS1 (col20) should have 交货汇总's 备注 (col13)
    assert str(row[20]) == "SUMMARY_BZ_VALUE"
    # 备注 (col21) should have 交货汇总's B_ADDRESS1 (col12)
    assert str(row[21]) == "SUMMARY_BA_VALUE"
    wb.close()


# =====================================================================
# RED 10: 已发运 not modified
# =====================================================================

def test_merge_yifayun_not_modified(tmp_path):
    """已发运 sheet should not be modified by merge."""
    master_path = str(tmp_path / "master.xlsx")
    mail_path = str(tmp_path / "mail.xlsx")
    output_path = str(tmp_path / "output.xlsx")

    yifayun_row = tuple([None] * 29)
    yifayun_row_list = list(yifayun_row)
    yifayun_row_list[5] = "2424827180"  # 订单号 in 已发运
    yifayun_row_list[9] = "test customer"
    yifayun_row = tuple(yifayun_row_list)

    _create_master_with_weifayun(
        master_path,
        detail_rows=[_make_master_row("2424827207")],
        weifayun_rows=[],
        yifayun_rows=[yifayun_row],
    )

    summary_rows = [_make_summary_row(jiaohuo="2424827180")]
    _create_mail_with_summary(mail_path, filter_rows=[], summary_rows=summary_rows)

    merge_mail_into_master(master_path, mail_path, output_path)

    wb = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
    ws = wb["已发运"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    data_rows = [r for r in rows if r[5] is not None]

    # Should still have only 1 row, unchanged
    assert len(data_rows) == 1
    assert str(data_rows[0][9]) == "test customer"
    wb.close()


# =====================================================================
# RED 11: Sheet5 not modified
# =====================================================================

def test_merge_sheet5_not_modified(tmp_path):
    """Sheet5 should not be modified by merge."""
    master_path = str(tmp_path / "master.xlsx")
    mail_path = str(tmp_path / "mail.xlsx")
    output_path = str(tmp_path / "output.xlsx")

    _create_master_file(master_path, [_make_master_row("2424827207")])
    _create_mail_file(mail_path, [_make_mail_row("2424827180")])

    # Add some data to Sheet5
    wb_m = openpyxl.load_workbook(master_path)
    ws5 = wb_m["Sheet5"]
    ws5.cell(row=1, column=1, value="2424787087")
    ws5.cell(row=1, column=3, value="HN79")
    ws5.cell(row=2, column=1, value="2424787217")
    ws5.cell(row=2, column=3, value="HN79")
    wb_m.save(master_path)
    wb_m.close()

    merge_mail_into_master(master_path, mail_path, output_path)

    wb = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
    ws5 = wb["Sheet5"]
    rows = list(ws5.iter_rows(min_row=1, values_only=True))

    assert rows[0][0] == "2424787087"
    assert rows[0][2] == "HN79"
    assert rows[1][0] == "2424787217"
    assert rows[1][2] == "HN79"
    wb.close()


# =====================================================================
# RED 12: 未发运 skips 交货号 already in 未发运
# =====================================================================

def test_merge_weifayun_skips_existing_in_weifayun(tmp_path):
    """交货号 already in 未发运 should not be added again."""
    master_path = str(tmp_path / "master.xlsx")
    mail_path = str(tmp_path / "mail.xlsx")
    output_path = str(tmp_path / "output.xlsx")

    existing_wfy = list([None] * 29)
    existing_wfy[5] = "2424827180"  # already in 未发运

    _create_master_with_weifayun(
        master_path,
        detail_rows=[_make_master_row("2424827207")],
        weifayun_rows=[tuple(existing_wfy)],
    )

    summary_rows = [_make_summary_row(jiaohuo="2424827180")]
    _create_mail_with_summary(mail_path, filter_rows=[], summary_rows=summary_rows)

    result = merge_mail_into_master(master_path, mail_path, output_path)

    wb = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
    ws = wb["未发运"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    data_rows = [r for r in rows if r[5] is not None]

    # Should still have only 1 row (no duplicate)
    assert len(data_rows) == 1
    assert result["appended_weifayun_count"] == 0
    wb.close()


# =====================================================================
# RED 13: 未发运 skips 交货号 already in 已发运
# =====================================================================

def test_merge_weifayun_skips_existing_in_yifayun(tmp_path):
    """交货号 already in 已发运 should not be added to 未发运."""
    master_path = str(tmp_path / "master.xlsx")
    mail_path = str(tmp_path / "mail.xlsx")
    output_path = str(tmp_path / "output.xlsx")

    yifayun_row = list([None] * 29)
    yifayun_row[5] = "2424827180"  # already in 已发运

    _create_master_with_weifayun(
        master_path,
        detail_rows=[_make_master_row("2424827207")],
        weifayun_rows=[],
        yifayun_rows=[tuple(yifayun_row)],
    )

    summary_rows = [_make_summary_row(jiaohuo="2424827180")]
    _create_mail_with_summary(mail_path, filter_rows=[], summary_rows=summary_rows)

    result = merge_mail_into_master(master_path, mail_path, output_path)

    wb = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
    ws = wb["未发运"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    data_rows = [r for r in rows if r[5] is not None]

    assert len(data_rows) == 0
    assert result["appended_weifayun_count"] == 0
    wb.close()


# =====================================================================
# RED 14: 未发运 skips 总计 row in 交货汇总
# =====================================================================

def test_merge_weifayun_skips_total_row(tmp_path):
    """交货汇总's 总计 row should not create a 未发运 entry."""
    master_path = str(tmp_path / "master.xlsx")
    mail_path = str(tmp_path / "mail.xlsx")
    output_path = str(tmp_path / "output.xlsx")

    _create_master_with_weifayun(
        master_path,
        detail_rows=[_make_master_row("2424827207")],
        weifayun_rows=[],
    )

    # One real row + one 总计 row (col0='总计', col5=None)
    total_row = tuple(["总计"] + [None] * 15)
    summary_rows = [
        _make_summary_row(jiaohuo="2424827180"),
        total_row,
    ]
    _create_mail_with_summary(mail_path, filter_rows=[], summary_rows=summary_rows)

    result = merge_mail_into_master(master_path, mail_path, output_path)

    wb = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
    ws = wb["未发运"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    data_rows = [r for r in rows if r[5] is not None]

    # Only 1 row (总计 skipped)
    assert len(data_rows) == 1
    assert result["appended_weifayun_count"] == 1
    wb.close()


# =====================================================================
# RED 15: 未发运 field mapping correct
# =====================================================================

def test_merge_weifayun_field_mapping(tmp_path):
    """交货汇总 fields should map to correct 未发运 columns."""
    master_path = str(tmp_path / "master.xlsx")
    mail_path = str(tmp_path / "mail.xlsx")
    output_path = str(tmp_path / "output.xlsx")

    _create_master_with_weifayun(
        master_path,
        detail_rows=[_make_master_row("2424827207")],
        weifayun_rows=[],
    )

    summary_rows = [
        _make_summary_row(
            jiaohuo="2424827180",
            bao_date=datetime.datetime(2026, 8, 22),
            huo_date=datetime.datetime(2026, 8, 28),
            city="汝城县",
            yundafang="18011338",
            xiaoshou="5507040770",
            name="朝批方盛",
            street="湖南省郴州市",
            qty=32,
            weight=0.16,
            volume=0.521,
            factory="801",
            b_address1="BA_VALUE",
            beizhu="BZ_VALUE",
        ),
    ]
    _create_mail_with_summary(mail_path, filter_rows=[], summary_rows=summary_rows)

    merge_mail_into_master(master_path, mail_path, output_path)

    wb = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
    ws = wb["未发运"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    data_rows = [r for r in rows if r[5] is not None]

    assert len(data_rows) == 1
    row = data_rows[0]
    # col0 下单日期 = 交货汇总 col0 发货日期
    assert row[0] == datetime.datetime(2026, 8, 22)
    # col1 需求日期 = 交货汇总 col1 交货日期
    assert row[1] == datetime.datetime(2026, 8, 28)
    # col2 到货城市 = 交货汇总 col2 送达方地点
    assert str(row[2]) == "汝城县"
    # col3 客户代码 = 交货汇总 col3 运达方
    assert str(row[3]) == "18011338"
    # col4 销售凭证 = 交货汇总 col4 销售凭证
    assert str(row[4]) == "5507040770"
    # col5 订单号 = 交货汇总 col5 交货
    assert str(row[5]) == "2424827180"
    # col9 客户名称 = 交货汇总 col6 运达方的名字
    assert str(row[9]) == "朝批方盛"
    # col10 客户地址 = 交货汇总 col7 街道
    assert str(row[10]) == "湖南省郴州市"
    # col11 数量 = 交货汇总 col8 求和项:交货量
    assert row[11] == 32
    # col12 吨位 = 交货汇总 col9 求和项:总重量
    assert row[12] == 0.16
    # col13 体积 = 交货汇总 col10 求和项:业务量
    assert row[13] == 0.521
    # col14 库区 = 交货汇总 col11 工厂
    assert str(row[14]) == "801"
    # col20 B_ADDRESS1 = 交货汇总 col13 备注 (swapped!)
    assert str(row[20]) == "BZ_VALUE"
    # col21 备注 = 交货汇总 col12 B_ADDRESS1 (swapped!)
    assert str(row[21]) == "BA_VALUE"
    wb.close()
